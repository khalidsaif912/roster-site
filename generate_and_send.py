import os
import re
import json
import calendar
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO

import requests
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText


# =========================
# Settings / Secrets
# =========================
EXCEL_URL = os.environ.get("EXCEL_URL", "").strip()

# Optional: a plain-text file containing the original roster filename
# (used to display the source name on the website).
SOURCE_NAME_URL = os.environ.get("SOURCE_NAME_URL", "").strip()
SOURCE_NAME_FALLBACK = os.environ.get("SOURCE_NAME_FALLBACK", "latest.xlsx").strip()


SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "").strip()
SMTP_PASS = os.environ.get("SMTP_PASS", "").strip()
MAIL_FROM = os.environ.get("MAIL_FROM", "").strip()
MAIL_TO = os.environ.get("MAIL_TO", "").strip()

PAGES_BASE_URL = os.environ.get("PAGES_BASE_URL", "").strip()  # optional
TZ = ZoneInfo("Asia/Muscat")
AUTO_OPEN_ACTIVE_SHIFT_IN_FULL = True

# Local cache directory inside repo (committed by actions)
ROSTERS_DIR = os.environ.get("ROSTERS_DIR", "rosters").strip() or "rosters"
# Excel sheets
DEPARTMENTS = [
    ("Officers", "Officers"),
    ("Supervisors", "Supervisors"),
    ("Load Control", "Load Control"),
    ("Export Checker", "Export Checker"),
    ("Export Operators", "Export Operators"),
    ("Unassigned", "Unassigned"),  # ← القسم الجديد
]

# For day-row matching only
DAYS = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

SHIFT_MAP = {
    "MN06": ("MN06", "Morning"),
    "ME06": ("ME06", "Morning"),
    "ME07": ("ME07", "Morning"),
    "MN12": ("MN12", "Afternoon"),
    "AN13": ("AN13", "Afternoon"),
    "AE14": ("AE14", "Afternoon"),
    "NN21": ("NN21", "Night"),
    "NE22": ("NE22", "Night"),
}

# تم تحويل كل الأسماء للإنجليزية
GROUP_ORDER = ["Morning", "Afternoon", "Night", "Standby", "Off Day", "Annual Leave", "Sick Leave", "Training", "Other"]


# =========================
# Helpers
# =========================
def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00A0", " ")).strip()

def to_western_digits(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    arabic = {'٠':'0','١':'1','٢':'2','٣':'3','٤':'4','٥':'5','٦':'6','٧':'7','٨':'8','٩':'9'}
    farsi  = {'۰':'0','۱':'1','۲':'2','۳':'3','۴':'4','۵':'5','۶':'6','۷':'7','۸':'8','۹':'9'}
    mp = {**arabic, **farsi}
    return "".join(mp.get(ch, ch) for ch in s)

def norm(s) -> str:
    return clean(to_western_digits(s))

def looks_like_time(s: str) -> bool:
    up = norm(s).upper()
    return bool(
        re.match(r"^\d{3,4}\s*H?\s*-\s*\d{3,4}\s*H?$", up)
        or re.match(r"^\d{3,4}\s*H$", up)
        or re.match(r"^\d{3,4}$", up)
    )

def looks_like_employee_name(s: str) -> bool:
    v = norm(s)
    if not v:
        return False
    up = v.upper()
    if looks_like_time(up):
        return False
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", up):
        return False
    # قوي: اسم - رقم
    if re.search(r"-\s*\d{3,}", v) and re.search(r"[A-Za-z\u0600-\u06FF]", v):
        return True
    # بديل: كلمتين أو أكثر
    parts = [p for p in v.split(" ") if p]
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", v) and len(parts) >= 2)

def looks_like_shift_code(s: str) -> bool:
    v = norm(s).upper()
    if not v:
        return False
    if looks_like_time(v):
        return False
    if v in ["OFF", "O", "LV", "TR", "ST", "SL", "AL", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14", "OT"]:
        return True
    if re.match(r"^(MN|AN|NN|NT|ME|AE|NE)\d{1,2}", v):
        return True
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", v):
        return True
    # ← إضافة: أي كود غريب مثل STAR14 (حرف+رقم) يعتبر shift code
    # مقيّد بوجود رقم واحد على الأقل لتجنب تصنيف الأسماء خطأً
    if len(v) >= 3 and re.search(r"[A-Z]", v) and re.search(r"\d", v):
        return True
    return False

def map_shift(code: str):
    c0 = norm(code)
    c = c0.upper()
    if not c or c == "0":
        return ("-", "Other")

    # ✅ Leave types (separated)
    if c == "AL" or c == "LV" or "ANNUAL LEAVE" in c:
        return ("AL", "Annual Leave")

    if c == "SL" or "SICK LEAVE" in c:
        return ("SL", "Sick Leave")

    # Training
    if c in ["TR"] or "TRAINING" in c:
        return ("TR", "Training")

    # 🔹 Standby - إظهار الكود الأصلي
    if c in ["ST", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14"] or "STANDBY" in c:
        return (c0, "Standby")

    if c == "OT" or c.startswith("OT"):
        return (c0, "Standby")

    if c in ["OFF", "O"] or re.search(r"(REST|OFF\s*DAY|REST\/OFF)", c):
        return ("OFF", "Off Day")

    if c in SHIFT_MAP:
        return SHIFT_MAP[c]

    return (c0, "Other")

def current_shift_key(now: datetime) -> str:
    # Morning  06:00–13:00
    # Afternoon 13:01–21:00
    # Night    21:01–05:59
    t = now.hour * 60 + now.minute
    if t >= 21 * 60 + 1 or t < 6 * 60:
        return "Night"
    if t >= 13 * 60 + 1:
        return "Afternoon"
    return "Morning"

def download_excel(url: str) -> bytes:
    """Download the Excel file bytes.

    Notes for OneDrive/SharePoint:
    - Share links often return an HTML preview unless `download=1` is present.
    - We also validate the response to avoid crashing openpyxl on non-xlsx content.
    """
    if not url:
        raise ValueError("EXCEL_URL is empty")

    # Force direct download for common OneDrive/SharePoint share links
    try:
        from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse

        u = urlparse(url)
        host = (u.netloc or "").lower()
        if ("onedrive.live.com" in host) or ("1drv.ms" in host) or ("sharepoint.com" in host):
            qs = dict(parse_qsl(u.query, keep_blank_values=True))
            if "download" not in qs:
                qs["download"] = "1"
                u = u._replace(query=urlencode(qs, doseq=True))
                url = urlunparse(u)
    except Exception:
        pass

    headers = {
        "User-Agent": "Mozilla/5.0 (GitHub Actions) roster-site",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
    }
    r = requests.get(url, headers=headers, timeout=60, allow_redirects=True)
    r.raise_for_status()

    ctype = (r.headers.get("Content-Type") or "").lower()
    data = r.content or b""

    # Basic validation: xlsx is a ZIP container and starts with PK
    if not data.startswith(b"PK"):
        # Help debugging: show a short hint
        hint = ""
        if "text/html" in ctype:
            hint = " (got HTML preview page; check OneDrive link/download=1)"
        elif ctype.startswith("image/"):
            hint = " (got an image; EXCEL_URL points to wrong file)"
        raise ValueError(f"Downloaded file is not a valid .xlsx (Content-Type: {ctype or 'unknown'}){hint}")

    return data


def download_text(url: str) -> str:
    """Download a small text file (e.g., source_name.txt)."""
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    return r.text.strip()

def get_source_name() -> str:
    """Return the original roster file name for display on the website."""
    if SOURCE_NAME_URL:
        try:
            name = download_text(SOURCE_NAME_URL)
            if name:
                return name
        except Exception:
            pass
    return SOURCE_NAME_FALLBACK or "latest.xlsx"

def infer_pages_base_url():
    return "https://khalidsaif912.github.io/roster-site"



# =========================
# Month parsing + local cache (repo)
# =========================
MONTH_NAME_TO_NUM = {
    "january": 1, "jan": 1,
    "february": 2, "feb": 2,
    "march": 3, "mar": 3,
    "april": 4, "apr": 4,
    "may": 5,
    "june": 6, "jun": 6,
    "july": 7, "jul": 7,
    "august": 8, "aug": 8,
    "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10,
    "november": 11, "nov": 11,
    "december": 12, "dec": 12,
}

def month_key_from_filename(name: str) -> str | None:
    """Extract YYYY-MM from roster attachment file name (e.g., 'February 2026')."""
    if not name:
        return None
    n = name.lower()
    n = re.sub(r"[\._\-]+", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    m = re.search(
        r"\b(january|jan|february|feb|march|mar|april|apr|may|june|jun|july|jul|august|aug|september|sep|sept|october|oct|november|nov|december|dec)\b\s+(\d{4})\b",
        n,
    )
    if not m:
        return None
    mon_name, year_s = m.group(1), m.group(2)
    mon = MONTH_NAME_TO_NUM.get(mon_name)
    if not mon:
        return None
    return f"{int(year_s):04d}-{mon:02d}"

def add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    y = year
    m = month + delta
    while m <= 0:
        y -= 1
        m += 12
    while m > 12:
        y += 1
        m -= 12
    return y, m

def cache_paths(month_key: str) -> tuple[str, str]:
    os.makedirs(ROSTERS_DIR, exist_ok=True)
    return (
        os.path.join(ROSTERS_DIR, f"{month_key}.xlsx"),
        os.path.join(ROSTERS_DIR, f"{month_key}.meta.json"),
    )

def write_bytes(path: str, data: bytes):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "wb") as f:
        f.write(data)

def read_json(path: str) -> dict | None:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def write_json(path: str, obj: dict):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)

def try_load_cached_workbook(month_key: str):
    xlsx_path, _ = cache_paths(month_key)
    if not os.path.exists(xlsx_path):
        return None
    try:
        with open(xlsx_path, "rb") as f:
            return load_workbook(BytesIO(f.read()), data_only=True)
    except Exception:
        return None

def cached_source_name(month_key: str) -> str:
    _, meta_path = cache_paths(month_key)
    meta = read_json(meta_path) or {}
    return (meta.get("original_filename") or meta.get("source_name") or "").strip()

# =========================
# Detect rows/cols (Days row + Date numbers row)
# =========================
def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    count = 0
    for d in DAYS:
        if any(d in x for x in ups):
            count += 1
    return count

def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False

def find_days_and_dates_rows(ws, scan_rows: int = 80):
    """
    يبحث عن صف فيه SUN..SAT بكثرة ثم صف تحته فيه أرقام 1..31
    """
    max_r = min(ws.max_row, scan_rows)
    days_row = None

    for r in range(1, max_r + 1):
        vals = _row_values(ws, r)
        if _count_day_tokens(vals) >= 3:
            days_row = r
            break

    if not days_row:
        return None, None

    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        vals = _row_values(ws, r)
        nums = sum(1 for v in vals if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break

    return days_row, date_row

def find_day_col(ws, days_row: int, date_row: int, today_dow: int, today_day: int):
    """
    يثبت العمود الصحيح باستخدام اليوم + رقم التاريخ
    """
    if not days_row or not date_row:
        return None

    day_key = DAYS[today_dow]
    # Prefer (day + date) match
    for c in range(1, ws.max_column + 1):
        top = norm(ws.cell(row=days_row, column=c).value).upper()
        bot = norm(ws.cell(row=date_row, column=c).value)
        if day_key in top and _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    # Fallback: date-only
    for c in range(1, ws.max_column + 1):
        bot = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    return None


def get_daynum_to_col(ws, date_row: int):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(v):
            m[int(float(v))] = c
    return m

def find_employee_col(ws, start_row: int):
    for c in range(1, min(ws.max_column, 15) + 1):
        found = 0
        for r in range(start_row, min(start_row + 20, ws.max_row) + 1):
            v = norm(ws.cell(row=r, column=c).value)
            if looks_like_employee_name(v):
                found += 1
        if found >= 3:
            return c
    return None

def range_suffix_for_day(day: int, daynum_to_raw: dict, code_key: str):
    """
    إذا كان يوم (day) جزء من block متصل من نفس code_key، يرجع (من X إلى Y)
    """
    sorted_days = sorted(daynum_to_raw.keys())
    if day not in sorted_days:
        return ""

    up_key = code_key.upper()

    # تحديد الأكواد المقبولة لهذا النوع من الإجازة/التدريب
    acceptable_codes = []
    if up_key in ["AL", "LV"] or "ANNUAL" in up_key:
        # الإجازة السنوية
        acceptable_codes = ["AL", "LV", "ANNUAL LEAVE"]
    elif up_key == "SL" or "SICK" in up_key:
        # الإجازة المرضية
        acceptable_codes = ["SL", "SICK LEAVE"]
    elif up_key == "TR" or "TRAINING" in up_key:
        # التدريب
        acceptable_codes = ["TR", "TRAINING"]
    else:
        # أي كود آخر - يجب أن يكون مطابق تماماً
        acceptable_codes = [up_key]

    def is_same_type(val: str) -> bool:
        """تحقق إذا كان الكود من نفس النوع"""
        if not val:
            return False
        val_upper = val.upper()
        for code in acceptable_codes:
            if code in val_upper or val_upper == code:
                return True
        return False

    # إيجاد بداية ونهاية النطاق المتصل
    start = day
    end = day
    
    # البحث للخلف لإيجاد بداية النطاق
    current = day - 1
    while current in sorted_days:
        val = norm(daynum_to_raw.get(current, ""))
        if is_same_type(val):
            start = current
            current -= 1
        else:
            break
    
    # البحث للأمام لإيجاد نهاية النطاق
    current = day + 1
    while current in sorted_days:
        val = norm(daynum_to_raw.get(current, ""))
        if is_same_type(val):
            end = current
            current += 1
        else:
            break

    if start == end:
        return ""
    return f"(<span style='font-size:0.75em;opacity:0.8;'>FROM</span> {start} <span style='font-size:0.75em;opacity:0.8;'>TO</span> {end})"




# =========================
# Department card colors
# =========================
DEPT_COLORS = [
    {"name": "blue",   "base": "#2563eb", "light": "#2563eb15", "border": "#2563eb18", "grad_from": "#2563eb", "grad_to": "#2563ebcc"},
    {"name": "cyan",   "base": "#0891b2", "light": "#0891b215", "border": "#0891b218", "grad_from": "#0891b2", "grad_to": "#0891b2cc"},
    {"name": "green",  "base": "#059669", "light": "#05966915", "border": "#05966918", "grad_from": "#059669", "grad_to": "#059669cc"},
    {"name": "red",    "base": "#dc2626", "light": "#dc262615", "border": "#dc262618", "grad_from": "#dc2626", "grad_to": "#dc2626cc"},
    {"name": "purple", "base": "#7c3aed", "light": "#7c3aed15", "border": "#7c3aed18", "grad_from": "#7c3aed", "grad_to": "#7c3aedcc"},
    {"name": "orange", "base": "#ea580c", "light": "#ea580c15", "border": "#ea580c18", "grad_from": "#ea580c", "grad_to": "#ea580ccc"},
]

# قسم Unassigned يأخذ لون برتقالي/رمادي
UNASSIGNED_COLOR = {"name": "gray", "base": "#6b7280", "light": "#6b728015", "border": "#6b728018", "grad_from": "#6b7280", "grad_to": "#6b7280cc"}

# =========================
# Shift group colors (Morning/Afternoon/Night/etc.)
# =========================
SHIFT_COLORS = {
    "Morning": {
        "border": "#f59e0b44",
        "bg": "#fef3c7",
        "summary_bg": "#fef3c7",
        "summary_border": "#f59e0b33",
        "label_color": "#92400e",
        "count_bg": "#f59e0b22",
        "count_color": "#92400e",
        "status_color": "#92400e",
        "icon": "☀️",
    },
    "Afternoon": {
        "border": "#f9731644",
        "bg": "#ffedd5",
        "summary_bg": "#ffedd5",
        "summary_border": "#f9731633",
        "label_color": "#9a3412",
        "count_bg": "#f9731622",
        "count_color": "#9a3412",
        "status_color": "#9a3412",
        "icon": "🌤️",
    },
    "Night": {
        "border": "#8b5cf644",
        "bg": "#ede9fe",
        "summary_bg": "#ede9fe",
        "summary_border": "#8b5cf633",
        "label_color": "#5b21b6",
        "count_bg": "#8b5cf622",
        "count_color": "#5b21b6",
        "status_color": "#5b21b6",
        "icon": "🌙",
    },
    "Off Day": {
        "border": "#6366f144",
        "bg": "#e0e7ff",
        "summary_bg": "#e0e7ff",
        "summary_border": "#6366f133",
        "label_color": "#3730a3",
        "count_bg": "#6366f122",
        "count_color": "#3730a3",
        "status_color": "#3730a3",
        "icon": "🛋️",
    },
    "Annual Leave": {
        "border": "#10b98144",
        "bg": "#d1fae5",
        "summary_bg": "#d1fae5",
        "summary_border": "#10b98133",
        "label_color": "#065f46",
        "count_bg": "#10b98122",
        "count_color": "#065f46",
        "status_color": "#065f46",
        "icon": "✈️",
    },
    "Training": {
        "border": "#0ea5e944",
        "bg": "#e0f2fe",
        "summary_bg": "#e0f2fe",
        "summary_border": "#0ea5e933",
        "label_color": "#075985",
        "count_bg": "#0ea5e922",
        "count_color": "#075985",
        "status_color": "#075985",
        "icon": "📚",
    },
    "Standby": {
        "border": "#9e9e9e44",
        "bg": "#f0f0f0",
        "summary_bg": "#f0f0f0",
        "summary_border": "#9e9e9e33",
        "label_color": "#555555",
        "count_bg": "#cccccc22",
        "count_color": "#555555",
        "status_color": "#555555",
        "icon": "🧍"
    }, 
    "Sick Leave": {
    "border": "#ef444444",
    "bg": "#fee2e2",
    "summary_bg": "#fee2e2",
    "summary_border": "#ef444433",
    "label_color": "#991b1b",
    "count_bg": "#ef444422",
    "count_color": "#991b1b",
    "status_color": "#991b1b",
    "icon": "🤒",
   },
    "Other": {
        "border": "#94a3b844",
        "bg": "#f1f5f9",
        "summary_bg": "#f1f5f9",
        "summary_border": "#94a3b833",
        "label_color": "#475569",
        "count_bg": "#94a3b822",
        "count_color": "#475569",
        "status_color": "#475569",
        "icon": "❓",
    },
}


# =========================
# HTML Builders
# =========================
def dept_card_html(dept_name: str, dept_color: dict, buckets: dict, open_group: str = None) -> str:
    # buckets = {group_key: [{"name": ..., "shift": ...}, ...], ...}
    total = sum(len(buckets.get(k, [])) for k in GROUP_ORDER)
    if total == 0:
        return ""

    shifts_html = ""
    for group_key in GROUP_ORDER:
        emps = buckets.get(group_key, [])
        if not emps:
            continue

        # Determine shift display name (use English directly)
        if group_key == "Morning":
            display_name = "Morning"
        elif group_key == "Afternoon":
            display_name = "Afternoon"
        elif group_key == "Night":
            display_name = "Night"
        elif group_key == "Off Day":
            display_name = "Off Day"
        elif group_key == "Annual Leave":
            display_name = "Annual Leave"
        elif group_key == "Sick Leave":
           display_name = "Sick Leave"
        elif group_key == "Training":
            display_name = "Training"
        elif group_key == "Standby":
            display_name = "Standby"
        else:
            display_name = "Other"

        colors = SHIFT_COLORS.get(group_key, SHIFT_COLORS["Other"])
        count = len(emps)
        open_attr = ' open' if (group_key == open_group) else ''

        rows_html = ""
        for i, e in enumerate(emps):
            alt = " empRowAlt" if i % 2 == 1 else ""
            rows_html += f"""<div class="empRow{alt}">
       <span class="empName">{e['name']}</span>
       <span class="empStatus" style="color:{colors['status_color']};">{e['shift']}</span>
     </div>"""

        shifts_html += f"""
    <details class="shiftCard" data-shift="{group_key}" style="border:1px solid {colors['border']}; background:{colors['bg']}"{open_attr}>
      <summary class="shiftSummary" data-shift-key="{group_key}" style="background:{colors['summary_bg']}; border-bottom:1px solid {colors['summary_border']};">
        <span class="shiftIcon">{colors['icon']}</span>
        <span class="shiftLabel" style="color:{colors['label_color']};">{display_name}</span>
        <span class="shiftCount" style="background:{colors['count_bg']}; color:{colors['count_color']};">{count}</span>
      </summary>
      <div class="shiftBody">
        {rows_html}
      </div>
    </details>
            """

    icon_svg = """
<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>
"""

    return f"""
    <div class="deptCard" data-dept="{dept_name}">
      <div style="height:5px; background:linear-gradient(to right, {dept_color['grad_from']}, {dept_color['grad_to']});"></div>

      <div class="deptHead" data-dept="{dept_name}" style="border-bottom:2px solid {dept_color['border']}; cursor:pointer; user-select:none; -webkit-user-select:none;">
        <div class="deptIcon" style="background:{dept_color['light']}; color:{dept_color['base']};">
          {icon_svg}
        </div>
        <div class="deptTitle">{dept_name}</div>
        <div class="deptBadge" style="background:{dept_color['light']}; color:{dept_color['base']}; border:1px solid {dept_color['border']};">
          <span style="font-size:10px;opacity:.7;display:block;margin-bottom:1px;text-transform:uppercase;letter-spacing:.5px;">Total</span>
          <span style="font-size:17px;font-weight:900;">{total}</span>
        </div>
      </div>

      <div class="shiftStack">
{shifts_html}
      </div>
    </div>
    """

def page_shell_html(date_label: str, iso_date: str, employees_total: int, departments_total: int,
                     dept_cards_html: str, cta_url: str, sent_time: str, source_name: str = "", last_updated: str = "", is_now_page: bool = False,
                     min_date: str = "", max_date: str = "", notice_html: str = "") -> str:

    # ⬅️ أضف هذا السطر
    pages_base = (PAGES_BASE_URL or infer_pages_base_url()).rstrip("/")
    min_attr = f'min="{min_date}"' if min_date else ""
    max_attr = f'max="{max_date}"' if max_date else ""

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@700;800;900&display=swap" rel="stylesheet">
  <style>
    /* ═══════ RESET ═══════ */
    body {{
      margin:0; padding:0;
      background:#eef1f7;
      font-family:'Segoe UI', system-ui, -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif;
      color:#0f172a;
      -webkit-font-smoothing:antialiased;
    }}
    * {{ box-sizing:border-box; }}

    /* ═══════ WRAP ═══════ */
    .wrap {{ max-width:680px; margin:0 auto; padding:16px 14px 28px; }}

    /* ═══════ HEADER ═══════ */
    .header {{
      background:linear-gradient(135deg, #1e40af 0%, #1976d2 50%, #0ea5e9 100%);
      color:#fff;
      padding:26px 18px 24px;
      border-radius:20px;
      text-align:center;
      box-shadow:0 8px 28px rgba(30,64,175,.25);
      position:relative;
      overflow:hidden;
    }}
    .header::before {{
      content:''; position:absolute;
      top:-30px; right:-40px;
      width:140px; height:140px;
      border-radius:50%;
      background:rgba(255,255,255,.08);
    }}
    .header::after {{
      content:''; position:absolute;
      bottom:-50px; left:-30px;
      width:160px; height:160px;
      border-radius:50%;
      background:rgba(255,255,255,.06);
    }}
    .header h1 {{ margin:0; font-size:24px; font-weight:800; position:relative; z-index:1; letter-spacing:-.3px; }}

    /* زر اللغة */
    .langToggle {{
      position:absolute; top:14px; right:16px; z-index:10;
      background:rgba(255,255,255,.18); border:2px solid rgba(255,255,255,.25);
      border-radius:50%; width:32px; height:32px;
      display:flex; align-items:center; justify-content:center;
      color:#fff; font-size:13px; font-weight:800; cursor:pointer;
      transition:all .25s; -webkit-tap-highlight-color:transparent; padding:0;
    }}
    .langToggle:hover {{ background:rgba(255,255,255,.30); transform:scale(1.08); }}
    body.ar {{ direction:rtl; font-family:'Segoe UI',Tahoma,Arial,sans-serif; }}
    .empRow, .empName, .empStatus {{ direction:ltr !important; unicode-bidi:embed; text-align:left !important; }}

    /* رسالة الترحيب */
    .welcomeMsg {{ display:none !important; }}

    /* ═══════ WELCOME CHIP ═══════ */
    .welcomeChip {{
      background:#fff; border:1px solid rgba(15,23,42,.1);
      border-radius:14px; padding:10px 14px;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
      transition:all .25s ease; min-width:80px;
      cursor:pointer; display:none;
      flex-direction:column; align-items:center; gap:1px;
    }}
    .welcomeChip.visible {{ display:flex; }}
    .welcomeChip:hover {{ transform:translateY(-3px); box-shadow:0 8px 20px rgba(99,102,241,.18); }}
    .welcomeChipIcon {{
      font-size:20px; display:block;
      animation:waveHand 2.8s ease-in-out infinite;
      transform-origin:70% 80%;
    }}
    @keyframes waveHand {{
      0%,55%,100% {{ transform:rotate(0deg); }}
      10%  {{ transform:rotate(20deg); }}
      20%  {{ transform:rotate(-10deg); }}
      30%  {{ transform:rotate(18deg); }}
      40%  {{ transform:rotate(-8deg); }}
      50%  {{ transform:rotate(12deg); }}
    }}
    .welcomeChipName {{
      font-size:13px; font-weight:900; color:#1e293b;
      font-family:'Outfit',system-ui,sans-serif;
      letter-spacing:-.4px; white-space:nowrap;
      max-width:90px; overflow:hidden; text-overflow:ellipsis;
      margin-top:1px;
    }}
    .welcomeChipLabel {{
      font-size:8.5px; font-weight:600; color:#64748b;
      text-transform:uppercase; letter-spacing:.5px; white-space:nowrap;
    }}

    /* ═══════ FLIP CHIP (موظفون ↔ أقسام) ═══════ */
    .flipChip {{
      background:#fff; border:1px solid rgba(15,23,42,.1);
      border-radius:14px; padding:10px 12px;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
      transition:box-shadow .25s,transform .25s;
      min-width:72px; position:relative;
      overflow:hidden; perspective:600px;
    }}
    .flipChip:hover {{ transform:translateY(-3px); box-shadow:0 8px 20px rgba(15,23,42,.12); }}
    .flipChipInner {{
      position:relative; width:100%; height:44px;
      transform-style:preserve-3d;
      transition:transform .7s cubic-bezier(.4,0,.2,1);
    }}
    .flipChipInner.flipped {{ transform:rotateX(180deg); }}
    .flipChipFace {{
      position:absolute; inset:0;
      display:flex; flex-direction:column;
      align-items:center; justify-content:center; gap:2px;
      backface-visibility:hidden; -webkit-backface-visibility:hidden;
    }}
    .flipChipBack {{ transform:rotateX(180deg); }}
    .flipChipVal {{
      font-size:22px; font-weight:900; color:#1e40af; line-height:1;
    }}
    .flipChipVal.green {{ color:#059669; }}
    .flipChipLabel {{
      font-size:9.5px; font-weight:600; color:#64748b;
      text-transform:uppercase; letter-spacing:.4px; white-space:nowrap;
    }}

    /* ═══════ CUSTOM DATE PICKER ═══════ */
    #datePicker {{
      position:absolute; width:1px; height:1px;
      opacity:0; pointer-events:none; font-size:16px;
      border:none; z-index:-1;
    }}
    .cdp-overlay {{
      display:none; position:fixed; inset:0;
      background:rgba(0,0,0,.4);
      backdrop-filter:blur(4px); -webkit-backdrop-filter:blur(4px);
      z-index:9998; align-items:flex-end; justify-content:center;
    }}
    .cdp-overlay.cdp-open {{ display:flex; animation:cdpFadeIn .18s ease; }}
    @keyframes cdpFadeIn {{ from{{opacity:0}} to{{opacity:1}} }}
    .cdp-sheet {{
      background:#fff; border-radius:22px 22px 0 0;
      width:100%; max-width:400px;
      box-shadow:0 -6px 32px rgba(15,23,42,.15);
      animation:cdpSlideUp .26s cubic-bezier(.22,1,.36,1);
      padding-bottom:max(16px,env(safe-area-inset-bottom));
    }}
    @keyframes cdpSlideUp {{
      from{{transform:translateY(100%);opacity:0}}
      to{{transform:translateY(0);opacity:1}}
    }}
    .cdp-handle {{
      width:36px; height:4px; background:#e2e8f0;
      border-radius:2px; margin:12px auto 0;
    }}
    .cdp-header {{
      display:flex; align-items:center;
      justify-content:space-between; padding:12px 16px 8px;
    }}
    .cdp-nav {{
      width:36px; height:36px; border-radius:10px;
      border:1.5px solid #e2e8f0; background:#f8fafc;
      font-size:18px; cursor:pointer; color:#475569;
      display:flex; align-items:center; justify-content:center;
      transition:all .12s; -webkit-tap-highlight-color:transparent;
    }}
    .cdp-nav:hover {{ background:#e2e8f0; color:#1e293b; }}
    .cdp-nav:active {{ transform:scale(.9); }}
    .cdp-nav:disabled {{ opacity:.25; cursor:default; pointer-events:none; }}
    .cdp-month-label {{
      font-size:15px; font-weight:700; color:#1e293b; letter-spacing:-.2px;
    }}
    .cdp-days-head {{
      display:grid; grid-template-columns:repeat(7,1fr);
      padding:0 10px; margin-bottom:2px;
    }}
    .cdp-days-head span {{
      text-align:center; font-size:10px; font-weight:700;
      color:#94a3b8; text-transform:uppercase;
      letter-spacing:.3px; padding:3px 0;
    }}
    .cdp-grid {{
      display:grid; grid-template-columns:repeat(7,1fr);
      gap:2px; padding:0 10px 14px;
    }}
    .cdp-day {{
      aspect-ratio:1; border-radius:9px; border:none;
      background:transparent; font-size:14px; font-weight:600;
      color:#334155; cursor:pointer;
      display:flex; align-items:center; justify-content:center;
      transition:background .1s,transform .1s;
      -webkit-tap-highlight-color:transparent; position:relative;
    }}
    .cdp-day:hover:not([disabled]):not(.cdp-sel) {{ background:#f1f5f9; }}
    .cdp-day:active:not([disabled]) {{ transform:scale(.85); }}
    .cdp-day[disabled] {{ opacity:.25; cursor:default; pointer-events:none; }}
    .cdp-day.cdp-today {{ color:#1e40af; font-weight:800; }}
    .cdp-day.cdp-today::after {{
      content:''; position:absolute; bottom:3px; left:50%;
      transform:translateX(-50%); width:4px; height:4px;
      background:#1e40af; border-radius:50%;
    }}
    .cdp-day.cdp-sel {{
      background:linear-gradient(135deg,#1e40af,#3b82f6);
      color:#fff; box-shadow:0 3px 10px rgba(30,64,175,.3);
    }}
    .cdp-day.cdp-sel.cdp-today::after {{ background:#fff; }}
    .cdp-empty {{ pointer-events:none; }}

    /* responsive */
    @media (max-width:600px) {{
      .welcomeChip {{ min-width:64px; padding:8px 10px; }}
      .welcomeChipName {{ font-size:11px; max-width:68px; }}
      .flipChip {{ min-width:60px; padding:8px 8px; }}
      .flipChipVal {{ font-size:18px; }}
      .flipChipLabel {{ font-size:8.5px; }}
      .flipChipInner {{ height:38px; }}
    }}
    .datePickerWrapper {{
      position:relative;
      display:inline-block;
      margin-top:14px;
      z-index:1;
    }}
    .header .dateTag {{
      display:inline-block;
      background:rgba(255,255,255,.18);
      padding:5px 18px;
      border-radius:10px;
      font-size:13px;
      font-weight:600;
      letter-spacing:.3px;
      cursor:pointer;
      transition:all .3s;
      border:2px solid rgba(255,255,255,.2);
      -webkit-tap-highlight-color:transparent;
      user-select:none;
      -webkit-user-select:none;
      direction:ltr;
    }}
    .header .dateTag:hover {{
      background:rgba(255,255,255,.25);
      transform:translateY(-1px);
    }}
    /* Date Picker Wrapper */


    /* ═══════ SUMMARY BAR ═══════ */
    .summaryBar {{ 
      display:flex; 
      justify-content:center; 
      align-items:stretch;
      gap:6px; 
      margin-top:14px;
      flex-wrap:wrap;
      padding:0 8px 2px;
      box-sizing:border-box;
      width:100%;
      max-width:100%;
      overflow:hidden;
    }}
    .summaryBar::-webkit-scrollbar {{ display:none; }}
    a.summaryChip:hover {{
      transform:translateY(-3px);
      box-shadow:0 8px 20px rgba(15,23,42,.12);
    }}
    a.summaryChip.gamesChip .chipVal {{ color:#7c3aed; }}
    a.summaryChip.gamesChip:hover {{ box-shadow:0 8px 20px rgba(124,58,237,.18); }}
    a.summaryChip.trainingChip .chipVal {{ color:#0ea5e9; }}
    a.summaryChip.trainingChip:hover {{ box-shadow:0 8px 20px rgba(14,165,233,.18); }}
    .summaryChip {{
      background:#fff;
      border:1px solid rgba(15,23,42,.1);
      border-radius:12px;
      padding:7px 8px;
      text-align:center;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
      transition:all .25s ease;
      min-width:56px;
      flex:1 1 auto;
      max-width:140px;
      box-sizing:border-box;
    }}
    .summaryChip .chipVal {{ font-size:18px; font-weight:900; color:#1e40af; line-height:1.2; }}
    .summaryChip .chipLabel {{ 
      font-size:8px;
      font-weight:600; 
      color:#64748b; 
      text-transform:uppercase; 
      letter-spacing:.3px; 
      margin-top:2px;
      white-space:nowrap;
    }}

    /* ═══════ SHIFT FILTER BUTTONS AS CHIPS ═══════ */
    button.summaryChip.shiftFilterBtn {{
      border:2px solid transparent;
      position:relative;
      overflow:hidden;
      padding:7px 8px;
    }}
    button.summaryChip.shiftFilterBtn:hover {{
      transform:translateY(-3px);
      box-shadow:0 8px 20px rgba(15,23,42,.12);
    }}
    button.summaryChip.shiftFilterBtn.active {{
      border-color:currentColor;
      box-shadow:0 6px 16px rgba(15,23,42,.18);
    }}
    button.summaryChip.shiftFilterBtn.active::before {{
      content:'';
      position:absolute;
      top:0;left:0;right:0;bottom:0;
      background:currentColor;
      opacity:.06;
    }}
    
    /* ألوان الورديات */
    button.shiftFilterBtn.morning {{ color:#f59e0b; }}
    button.shiftFilterBtn.morning .chipVal {{ color:#f59e0b; }}
    button.shiftFilterBtn.morning .chipLabel {{ color:#92400e; }}
    
    button.shiftFilterBtn.afternoon {{ color:#f97316; }}
    button.shiftFilterBtn.afternoon .chipVal {{ color:#f97316; }}
    button.shiftFilterBtn.afternoon .chipLabel {{ color:#9a3412; }}
    
    button.shiftFilterBtn.night {{ color:#8b5cf6; }}
    button.shiftFilterBtn.night .chipVal {{ color:#8b5cf6; }}
    button.shiftFilterBtn.night .chipLabel {{ color:#5b21b6; }}
    
    button.shiftFilterBtn.all {{ color:#1e40af; }}
    button.shiftFilterBtn.all .chipVal {{ color:#1e40af; }}
    button.shiftFilterBtn.all .chipLabel {{ color:#1e40af; }}

    /* الفليب شيب */
    .flipChip {{ min-width:0; flex-shrink:0; }}
    .flipChipInner {{ height:38px; }}
    .flipChipVal {{ font-size:18px; }}
    .flipChipLabel {{ font-size:8px; }}
    .welcomeChip {{ min-width:0; flex-shrink:0; padding:7px 8px; }}
    .welcomeChipName {{ font-size:11px; max-width:60px; }}
    .welcomeChipLabel {{ font-size:7.5px; }}

    @media (max-width:900px){{
      .summaryBar {{ justify-content:flex-start; }}
    }}


    /* ═══════ DEPARTMENT CARD ═══════ */
    .deptCard {{
      margin-top:18px;
      background:#fff;
      border-radius:18px;
      overflow:hidden;
      border:1px solid rgba(15,23,42,.07);
      box-shadow:0 4px 18px rgba(15,23,42,.08);
    }}
    .deptHead {{
      display:flex;
      align-items:center;
      gap:12px;
      padding:14px 16px;
      background:#fff;
      -webkit-touch-callout:none;
      -webkit-user-select:none;
      user-select:none;
      touch-action:manipulation;
    }}
    .deptIcon {{
      width:40px; height:40px;
      border-radius:12px;
      display:flex; align-items:center; justify-content:center;
      flex-shrink:0;
    }}
    .deptTitle {{ font-size:18px; font-weight:800; color:#1e293b; flex:1; letter-spacing:-.2px; }}
    .deptBadge {{ min-width:48px; padding:6px 10px; border-radius:12px; text-align:center; }}

    /* ═══════ SHIFT STACK ═══════ */
    .shiftStack {{ padding:10px; display:flex; flex-direction:column; gap:8px; }}

    /* ═══════ SHIFT CARD — <details> ═══════ */
    .shiftCard {{
      border-radius:14px;
      overflow:hidden;
    }}

    .shiftSummary {{
      display:flex;
      align-items:center;
      gap:10px;
      padding:11px 14px;
      cursor:pointer;
      list-style:none;
      -webkit-appearance:none;
      appearance:none;
      -webkit-touch-callout:none;
      -webkit-user-select:none;
      user-select:none;
      touch-action:manipulation;
    }}
    .shiftSummary::-webkit-details-marker {{ display:none; }}
    .shiftSummary::marker              {{ display:none; }}

    .shiftIcon  {{ font-size:20px; line-height:1; flex-shrink:0; }}
    .shiftLabel {{ font-size:15px; font-weight:800; flex:1; letter-spacing:-.1px; }}
    .shiftCount {{
      font-size:13px; font-weight:800;
      padding:3px 10px; border-radius:20px;
      flex-shrink:0;
    }}

    /* chevron يدور لما يفتح */
    .shiftSummary::after {{
      content:'▾';
      font-size:14px;
      color:#94a3b8;
      transition:transform .2s;
      flex-shrink:0;
    }}
    .shiftCard[open] .shiftSummary::after {{
      transform:rotate(180deg);
    }}

    .shiftBody {{ background:rgba(255,255,255,.7); }}

    /* ── employee row ── */
    .empRow {{
      display:flex;
      align-items:center;
      justify-content:space-between;
      padding:9px 16px;
      border-top:1px solid rgba(15,23,42,.06);
    }}
    .empRowAlt {{ background:rgba(15,23,42,.02); }}
    .empName  {{ font-size:15px; font-weight:700; color:#1e293b; }}
    .empStatus {{ font-size:13px; font-weight:600; }}

    /* ═══════ CTA ═══════ */
    .btnWrap {{ margin-top:20px; text-align:center; }}
    .btn {{
      display:inline-block;
      padding:14px 38px;
      border-radius:16px;
      background:linear-gradient(135deg, #1e40af, #1976d2);
      color:#fff !important;
      text-decoration:none;
      font-weight:800;
      font-size:15px;
      box-shadow:0 6px 20px rgba(30,64,175,.3);
    }}

    /* ═══════ FOOTER ═══════ */
    .footer {{ margin-top:18px; text-align:center; font-size:12px; color:#94a3b8; padding:12px 0; line-height:1.9; }}
    .footer strong {{ color:#64748b; }}

    /* ═══════ MOBILE ═══════ */
    @media (max-width:480px){{
      .wrap            {{ padding:12px 10px 22px; }}
      .header h1       {{ font-size:21px; }}
      .deptTitle       {{ font-size:16px; }}
      .empName         {{ font-size:14px; }}
      .empStatus       {{ font-size:12px; }}
      .shiftLabel      {{ font-size:14px; }}
      .summaryBar      {{ gap:8px; }}
      .summaryChip     {{ padding:8px 14px; }}
      .summaryChip .chipVal {{ font-size:19px; }}
    }}

  </style>
  <!-- html2canvas for screenshots -->
  <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
  <!-- Google Analytics - Export Page Tracking -->
  <script async src="https://www.googletagmanager.com/gtag/js?id=G-RQVWSXQVB5"></script>
  <script>
    window.dataLayer = window.dataLayer || [];
    function gtag(){{dataLayer.push(arguments);}}
    gtag('js', new Date());
    gtag('config', 'G-RQVWSXQVB5');
    gtag('event', 'page_view', {{
      'page_title': 'Export Duty Roster',
      'page_section': 'export',
      'page_date': '{iso_date}'
    }});
  </script>
</head>
<body>
<div class="wrap">

  <!-- ════ HEADER ════ -->
  <div class="header">
    <button class="langToggle" id="langToggle" onclick="toggleLang()">ع</button>
    <div class="welcomeMsg" id="welcomeMsg"></div>
    <h1 id="pageTitle">📋 Duty Roster</h1>
    <div class="datePickerWrapper">
      <button class="dateTag" id="dateTag" type="button" onclick="openDatePicker()">📅 {date_label}</button>
      <input id="datePicker" type="date" value="{iso_date}" {min_attr} {max_attr} />
    </div>
  </div>

  <!-- ═══ CUSTOM DATE PICKER OVERLAY ═══ -->
  <div class="cdp-overlay" id="cdpOverlay" onclick="cdpBgClick(event)">
    <div class="cdp-sheet">
      <div class="cdp-handle"></div>
      <div class="cdp-header">
        <button class="cdp-nav" id="cdpPrev" onclick="cdpMove(-1)" type="button">‹</button>
        <span class="cdp-month-label" id="cdpLabel"></span>
        <button class="cdp-nav" id="cdpNext" onclick="cdpMove(1)"  type="button">›</button>
      </div>
      <div class="cdp-days-head">
        <span>Su</span><span>Mo</span><span>Tu</span><span>We</span>
        <span>Th</span><span>Fr</span><span>Sa</span>
      </div>
      <div class="cdp-grid" id="cdpGrid"></div>
    </div>
  </div>

  {notice_html if notice_html else ""}

  <!-- ════ SUMMARY CHIPS ════ -->
  <div class="summaryBar">

    <!-- Welcome Chip -->
    <div class="welcomeChip" id="welcomeChip" onclick="goToMySchedule()" title="انقر للذهاب لجدولك">
      <span class="welcomeChipIcon">👋</span>
      <span class="welcomeChipName" id="welcomeChipName"></span>
      <span class="welcomeChipLabel" data-key="welcome">Hello</span>
    </div>

    <!-- Flip Chip: موظفون ↔ أقسام -->
    <div class="flipChip">
      <div class="flipChipInner" id="flipChipInner">
        <div class="flipChipFace">
          <div class="flipChipVal">{employees_total}</div>
          <div class="flipChipLabel" data-key="employees">Employees</div>
        </div>
        <div class="flipChipFace flipChipBack">
          <div class="flipChipVal green">{departments_total}</div>
          <div class="flipChipLabel" data-key="departments">Depts</div>
        </div>
      </div>
    </div>

    <a href="https://khalidsaif912.github.io/roster-site/my-schedules/index.html" id="myScheduleBtn" class="summaryChip" style="cursor:pointer;text-decoration:none;" onclick="goToMySchedule(event)">
      <div class="chipVal">🗓️</div>
      <div class="chipLabel" data-key="mySchedule">My Schedule</div>
    </a>
    <a href="https://dgr-exp.netlify.app" class="summaryChip gamesChip" style="cursor:pointer;text-decoration:none;">
      <div class="chipVal">🎮</div>
      <div class="chipLabel" data-key="games">Games</div>
    </a>
    <a href="{pages_base}/training/" class="summaryChip trainingChip" style="cursor:pointer;text-decoration:none;">
      <div class="chipVal">📚</div>
      <div class="chipLabel" data-key="training">Training</div>
    </a>
    {"" if not is_now_page else '''
    <button class="summaryChip shiftFilterBtn morning" data-shift="Morning" style="cursor:pointer;">
      <div class="chipVal">☀️</div>
      <div class="chipLabel" data-key="morning">Morning</div>
    </button>
    <button class="summaryChip shiftFilterBtn afternoon" data-shift="Afternoon" style="cursor:pointer;">
      <div class="chipVal">🌤️</div>
      <div class="chipLabel" data-key="afternoon">Afternoon</div>
    </button>
    <button class="summaryChip shiftFilterBtn night" data-shift="Night" style="cursor:pointer;">
      <div class="chipVal">🌙</div>
      <div class="chipLabel" data-key="night">Night</div>
    </button>
    <button class="summaryChip shiftFilterBtn all active" data-shift="All" style="cursor:pointer;">
      <div class="chipVal">📋</div>
      <div class="chipLabel" data-key="allShifts">All Shifts</div>
    </button>
    '''}
  </div>

  <!-- ════ DEPARTMENT CARDS ════ -->
  {dept_cards_html}

  <!-- ════ CTA ════ -->
  <div class="btnWrap">
    <a class="btn" id="ctaBtn" href="{cta_url}">📋 View Full Duty Roster</a>
  </div>

<div class="btnWrap">
  <a href="{pages_base}/subscribe/" class="btn" id="subscribeBtn">📩 Subscribe</a>
</div>

  <!-- ════ FOOTER ════ -->
  <div class="footer">
    <strong style="color:#475569;font-size:13px;" id="lastUpdatedTrigger">Last Updated:</strong> <strong style="color:#1e40af;">{last_updated}</strong>
    <br>
    <span style="font-size:11px;color:#94a3b8;display:inline-block;margin-top:4px;">
      📄 <strong style="color:#64748b;">{source_name}</strong>
    </span>
  </div>
</div>

<script>
(function(){{
  var picker = document.getElementById('datePicker');
  if(!picker) return;

  // ═══════════════════════════════════════════════════
  // التحقق من التاريخ وإعادة التوجيه للـ today
  // ═══════════════════════════════════════════════════
  function checkAndRedirectToToday() {{
    var path = window.location.pathname || '/';
    var isNowPage = path.includes('/now');

    var dateMatch = path.match(/\/date\/(\d{{4}})-(\d{{2}})-(\d{{2}})\//);
    if (dateMatch) {{
      var pageDate = dateMatch[1] + '-' + dateMatch[2] + '-' + dateMatch[3];

      var now = new Date();
      var muscatTime = new Date(now.getTime() + (4 * 60 * 60 * 1000) + (now.getTimezoneOffset() * 60 * 1000));
      var todayStr = muscatTime.getFullYear() + '-' +
        String(muscatTime.getMonth() + 1).padStart(2, '0') + '-' +
        String(muscatTime.getDate()).padStart(2, '0');

      if (pageDate !== todayStr) {{
        var isPageLoad = sessionStorage.getItem('pageLoaded');
        if (isPageLoad) {{
          sessionStorage.removeItem('pageLoaded');
          var basePath = path
            .replace(/\/date\/\d{{4}}-\d{{2}}-\d{{2}}\/.*$/, '/')
            .replace(/\/now\/.*$/, '/')
            .replace(/\/+$/, '');
          window.location.href = basePath + '/date/' + todayStr + '/' + (isNowPage ? 'now/' : '');
          return true;
        }} else {{
          sessionStorage.setItem('pageLoaded', 'true');
        }}
      }}
    }}
    return false;
  }}

  if (checkAndRedirectToToday()) return;

  // ═══════════════════════════════════════════════════
  // عند تغيير التاريخ → انتقل للصفحة المناسبة
  // ═══════════════════════════════════════════════════
  picker.addEventListener('change', function() {{
    if (!picker.value) return;

    sessionStorage.removeItem('pageLoaded');

    var path = window.location.pathname || '/';
    var isNowPage = path.includes('/now');
    var base = path
      .replace(/\/date\/\d{{4}}-\d{{2}}-\d{{2}}\/.*$/, '/')
      .replace(/\/now\/.*$/, '/')
      .replace(/\/+$/, '');

    var target = base + '/date/' + picker.value + '/';
    if (isNowPage) target += 'now/';

    window.location.href = target;
  }});

}})();

// ═══════════════════════════════════════════════════
// فتح الـ date picker - يعمل على iOS و Android و Desktop
// ═══════════════════════════════════════════════════
function goToMySchedule(e) {{
  if (e && e.preventDefault) e.preventDefault();
  var id = localStorage.getItem('savedEmpId');
  var base = 'https://khalidsaif912.github.io/roster-site/my-schedules/index.html';
  location.href = id ? base + '?emp=' + encodeURIComponent(id) : base;
}}
// ═══════════════════════════════════════════════════
(function(){{
  var filterBtns = document.querySelectorAll('.shiftFilterBtn');
  if(!filterBtns.length) return; // Not a /now/ page
  
  var allShiftCards = document.querySelectorAll('.shiftCard');
  
  // Group shift cards by shift type
  var shiftGroups = {{}};
  allShiftCards.forEach(function(card){{
    var shiftType = card.dataset.shift;
    if(!shiftType) return;
    if(!shiftGroups[shiftType]) shiftGroups[shiftType] = [];
    shiftGroups[shiftType].push(card);
  }});
  
  // Determine current shift based on Muscat time (Asia/Muscat = UTC+4)
  function getCurrentShift(){{
    var now = new Date();
    var muscatHour = parseInt(now.toLocaleString('en-US', {{timeZone:'Asia/Muscat', hour:'2-digit', hour12:false}}));
    var muscatMinute = parseInt(now.toLocaleString('en-US', {{timeZone:'Asia/Muscat', minute:'2-digit'}}));
    var t = muscatHour * 60 + muscatMinute;
    
    if(t >= 21 * 60 + 1 || t < 6 * 60) return 'Night';
    if(t >= 13 * 60 + 1) return 'Afternoon';
    return 'Morning';
  }}
  
  // Set active shift on load - default to current shift
  var currentShift = getCurrentShift();
  filterBtns.forEach(function(btn){{
    if(btn.dataset.shift === currentShift){{
      btn.classList.add('active');
    }} else {{
      btn.classList.remove('active');
    }}
  }});
  
  // Filter function
  function filterShifts(selectedShift){{
    var totalEmployees = 0;
    
    if(selectedShift === 'All'){{
      // Show all shifts
      allShiftCards.forEach(function(card){{
        card.style.display = '';
        var count = card.querySelector('.shiftCount');
        if(count) totalEmployees += parseInt(count.textContent) || 0;
      }});
    }} else {{
      // Hide all cards first
      allShiftCards.forEach(function(card){{ card.style.display = 'none'; }});
      
      // Show only selected shift cards and count employees
      if(shiftGroups[selectedShift]){{
        shiftGroups[selectedShift].forEach(function(card){{
          card.style.display = '';
          // Auto-open the selected shift
          card.setAttribute('open', '');
          // Count employees in this card
          var count = card.querySelector('.shiftCount');
          if(count) totalEmployees += parseInt(count.textContent) || 0;
        }});
      }}
      
      // Also show Off Day, Leave, Training, Standby in all shifts
      var alwaysShow = ['Off Day', 'Annual Leave', 'Sick Leave', 'Training', 'Standby', 'Other'];
      alwaysShow.forEach(function(type){{
        if(shiftGroups[type]){{
          shiftGroups[type].forEach(function(card){{
            card.style.display = '';
            // لا تحسب Off Day/Leave في عداد الفترة المختارة
          }});
        }}
      }});
    }}
    
    // Update employee count in summary
    var employeeChip = document.querySelector('.summaryChip .chipVal');
    if(employeeChip){{
      employeeChip.textContent = totalEmployees;
    }}
    
    // Update button states
    filterBtns.forEach(function(btn){{
      if(btn.dataset.shift === selectedShift){{
        btn.classList.add('active');
      }} else {{
        btn.classList.remove('active');
      }}
    }});
  }}
  
  // Add click handlers
  filterBtns.forEach(function(btn){{
    btn.addEventListener('click', function(){{
      filterShifts(this.dataset.shift);
    }});
  }});
  
  // Auto-filter on page load - show current shift
  filterShifts(currentShift);
}})();

// ══════════════════════════════════════════════════
// Language Toggle
// ══════════════════════════════════════════════════
var LANG = localStorage.getItem('rosterLang') || 'en';
var T = {{
  en: {{
    title:'📋 Duty Roster', langBtn:'ع',
    employees:'Emp.', departments:'Depts.', total:'Total',
    morning:'Morning', afternoon:'Afternoon', night:'Night',
    offday:'Off Day', annualLeave:'Annual Leave', sickLeave:'Sick Leave',
    training:'Training', standby:'Standby', other:'Other',
    from:'FROM', to:'TO',
    viewFull:'📋 View Full Duty Roster', subscribe:'📩 Subscribe',
    officers:'Officers', supervisors:'Supervisors', loadControl:'Load Control',
    exportChecker:'Export Checker', exportOps:'Export Operators', unassigned:'Unassigned',
    morning2:'Morning', afternoon2:'Afternoon', night2:'Night', allShifts:'All Shifts', mySchedule:'Schedule', games:'Games',
  }},
  ar: {{
    title:'📋 جدول المناوبات', langBtn:'EN',
    employees:'الموظفون', departments:'الأقسام', total:'المجموع',
    morning:'صباح', afternoon:'ظهر', night:'ليل',
    offday:'إجازة', annualLeave:'إجازة سنوية', sickLeave:'إجازة مرضية',
    training:'تدريب', standby:'احتياط', other:'أخرى',
    from:'من', to:'إلى',
    viewFull:'📋 عرض جدول المناوبات الكامل', subscribe:'📩 اشتراك',
    officers:'الضباط', supervisors:'المشرفون', loadControl:'مراقبة الحمولة',
    exportChecker:'مدقق الصادرات', exportOps:'مشغلو الصادرات', unassigned:'غير مُعيَّن',
    morning2:'صباح', afternoon2:'ظهر', night2:'ليل', allShifts:'الكل', mySchedule:'جدولي', games:'الألعاب',
  }}
}};

function applyLang(lang) {{
  var t=T[lang], isAr=lang==='ar';
  document.body.classList.toggle('ar',isAr);
  document.documentElement.setAttribute('lang',lang);
  var el=document.getElementById('pageTitle'); if(el) el.textContent=t.title;
  var btn=document.getElementById('langToggle'); if(btn) btn.textContent=t.langBtn;
  document.querySelectorAll('.chipLabel').forEach(function(el) {{
    var k=el.dataset.key;
    if(k==='employees') el.textContent=t.employees;
    else if(k==='departments') el.textContent=t.departments;
    else if(k==='morning') el.textContent=t.morning2;
    else if(k==='afternoon') el.textContent=t.afternoon2;
    else if(k==='night') el.textContent=t.night2;
    else if(k==='allShifts') el.textContent=t.allShifts;
    else if(k==='mySchedule') el.textContent=t.mySchedule;
  }});
  document.querySelectorAll('.deptBadge span:first-child').forEach(function(el) {{ el.textContent=t.total; }});
  var deptMap={{'Officers':t.officers,'Supervisors':t.supervisors,'Load Control':t.loadControl,
    'Export Checker':t.exportChecker,'Export Operators':t.exportOps,'Unassigned':t.unassigned}};
  document.querySelectorAll('.deptTitle').forEach(function(el) {{
    if(!el.dataset.key) el.dataset.key=el.textContent.trim();
    if(deptMap[el.dataset.key]) el.textContent=deptMap[el.dataset.key];
  }});
  var shiftMap={{'Morning':t.morning,'Afternoon':t.afternoon,'Night':t.night,
    'Off Day':t.offday,'Annual Leave':t.annualLeave,'Sick Leave':t.sickLeave,
    'Training':t.training,'Standby':t.standby,'Other':t.other}};
  document.querySelectorAll('.shiftLabel').forEach(function(el) {{
    if(!el.dataset.key) el.dataset.key=el.textContent.trim();
    if(shiftMap[el.dataset.key]) el.textContent=shiftMap[el.dataset.key];
  }});
  document.querySelectorAll('.empStatus span').forEach(function(el) {{
    var txt=el.textContent.trim();
    if(txt==='FROM'||txt==='من') el.textContent=t.from;
    if(txt==='TO'||txt==='إلى') el.textContent=t.to;
  }});
  var c1=document.getElementById('ctaBtn'); if(c1) c1.textContent=t.viewFull;
  var c2=document.getElementById('subscribeBtn'); if(c2) c2.textContent=t.subscribe;
  var footer=document.querySelector('.footer');
  if(footer) {{
    var h=footer.innerHTML;
    if(isAr) {{
      h=h.replace('Last Updated','آخر تحديث'); h=h.replace('Total:','المجموع:');
      h=h.replace(' employees',' موظف'); h=h.replace('Source:','المصدر:');
    }} else {{
      h=h.replace('آخر تحديث','Last Updated'); h=h.replace('المجموع:','Total:');
      h=h.replace(' موظف',' employees'); h=h.replace('المصدر:','Source:');
    }}
    footer.innerHTML=h;
  }}
  localStorage.setItem('rosterLang',lang);
  LANG=lang;
}}
function toggleLang() {{ applyLang(LANG==='en'?'ar':'en'); }}
applyLang(LANG);

// ═══════════════════════════════════════════════════
// FLIP CHIP + WELCOME CHIP + CUSTOM DATE PICKER
// ═══════════════════════════════════════════════════
(function() {{

  // ── 1. Flip Chip: يتبدل كل 3 ثوانٍ ──
  var inner = document.getElementById('flipChipInner');
  if (inner) setInterval(function() {{ inner.classList.toggle('flipped'); }}, 3000);

  // ── 2. Welcome Chip ──
  var empId = localStorage.getItem('savedEmpId');
  if (empId) {{
    var origin = location.origin;
    var base = location.pathname.includes('/roster-site/') ? origin+'/roster-site/' : origin+'/';
    fetch(base+'schedules/'+empId+'.json')
      .then(function(r){{ return r.ok?r.json():null; }})
      .then(function(d){{
        if(!d||!d.name) return;
        var chip=document.getElementById('welcomeChip');
        var nameEl=document.getElementById('welcomeChipName');
        if(chip&&nameEl){{
          nameEl.textContent=d.name.split(' ')[0];
          chip.classList.add('visible');
        }}
      }}).catch(function(){{}});
  }}

  // ── 3. Custom Date Picker ──
  var MIN_D='{min_date}', MAX_D='{max_date}', SEL_D='{iso_date}';
  var cy, cm;
  var MN=['January','February','March','April','May','June',
          'July','August','September','October','November','December'];

  function pad(n)   {{ return String(n).padStart(2,'0'); }}
  function ds(y,m,d){{ return y+'-'+pad(m)+'-'+pad(d); }}
  function today()  {{
    var n=new Date(),ms=new Date(n.getTime()+(4*3600000)+(n.getTimezoneOffset()*60000));
    return ds(ms.getFullYear(),ms.getMonth()+1,ms.getDate());
  }}

  function render() {{
    document.getElementById('cdpLabel').textContent=MN[cm-1]+' '+cy;
    var pv=document.getElementById('cdpPrev'), nx=document.getElementById('cdpNext');
    var cur=cy+'-'+pad(cm);
    pv.disabled = !!(MIN_D && cur<=MIN_D.slice(0,7));
    nx.disabled = !!(MAX_D && cur>=MAX_D.slice(0,7));
    var td=today(), html='';
    var fd=new Date(cy,cm-1,1).getDay();
    var dim=new Date(cy,cm,0).getDate();
    for(var i=0;i<fd;i++) html+='<div class="cdp-day cdp-empty"></div>';
    for(var d=1;d<=dim;d++) {{
      var dstr=ds(cy,cm,d);
      var dis=(MIN_D&&dstr<MIN_D)||(MAX_D&&dstr>MAX_D);
      var cls='cdp-day'+(dstr===td?' cdp-today':'')+(dstr===SEL_D?' cdp-sel':'');
      html+='<button class="'+cls+'" type="button" data-date="'+dstr+'"'+(dis?' disabled':'')+'>'+d+'</button>';
    }}
    var grid=document.getElementById('cdpGrid');
    grid.innerHTML=html;
    grid.onclick=function(e){{
      var btn=e.target;
      while(btn&&btn.tagName!=='BUTTON') btn=btn.parentElement;
      if(btn&&btn.dataset.date&&!btn.disabled) window.cdpPick(btn.dataset.date);
    }};
  }}

  window.cdpMove=function(d){{
    cm+=d; if(cm>12){{cm=1;cy++;}} if(cm<1){{cm=12;cy--;}} render();
  }};
  window.cdpPick=function(dstr){{
    document.getElementById('cdpOverlay').classList.remove('cdp-open');
    document.body.style.overflow='';
    sessionStorage.removeItem('pageLoaded');
    var path=window.location.pathname||'/';
    var isNow=path.includes('/now');
    var base=path.replace(/\/date\/\d{{4}}-\d{{2}}-\d{{2}}\/.*$/,'/')
                 .replace(/\/now\/.*$/,'/').replace(/\/+$/,'');
    window.location.href=base+'/date/'+dstr+'/'+(isNow?'now/':'');
  }};
  window.openDatePicker=function(){{
    var p=SEL_D||today(), pts=p.split('-');
    cy=+pts[0]; cm=+pts[1]; render();
    document.getElementById('cdpOverlay').classList.add('cdp-open');
    document.body.style.overflow='hidden';
  }};
  window.cdpBgClick=function(e){{
    if(e.target===document.getElementById('cdpOverlay')){{
      document.getElementById('cdpOverlay').classList.remove('cdp-open');
      document.body.style.overflow='';
    }}
  }};
  document.addEventListener('keydown',function(e){{
    if(e.key==='Escape'){{
      document.getElementById('cdpOverlay').classList.remove('cdp-open');
      document.body.style.overflow='';
    }}
  }});

}})();

// ✅ فتح المناوبة الصحيحة في الصفحة الرئيسية بناءً على توقيت مسقط
(function() {{
  var now = new Date();
  var muscatHour    = parseInt(now.toLocaleString('en-US', {{timeZone:'Asia/Muscat', hour:'2-digit', hour12:false}}));
  var muscatMinute  = parseInt(now.toLocaleString('en-US', {{timeZone:'Asia/Muscat', minute:'2-digit'}}));
  var t = muscatHour * 60 + muscatMinute;
  var activeShift = (t >= 21*60+1 || t < 6*60) ? 'Night' : (t >= 13*60+1) ? 'Afternoon' : 'Morning';
  document.querySelectorAll('details.shiftCard').forEach(function(el) {{
    el.removeAttribute('open');
  }});
  document.querySelectorAll('details.shiftCard[data-shift="' + activeShift + '"]').forEach(function(el) {{
    el.setAttribute('open', '');
  }});
}})();

// ✅ تسجيل الزيارة
(function() {{
  var SUPA_URL = 'https://tqfoiliyhlhidrrjfczq.supabase.co';
  var SUPA_KEY = 'sb_publishable_TQCEVXnTOEzoQsYAVVxidA_gUCx7Uez';

  var visitorId = localStorage.getItem('_vid');
  if(!visitorId) {{
    visitorId = 'v_' + Math.random().toString(36).slice(2) + Date.now().toString(36);
    localStorage.setItem('_vid', visitorId);
  }}

  var now = new Date();
  var muscat = new Date(now.getTime() + (4*60*60*1000) + (now.getTimezoneOffset()*60*1000));
  var todayStr = muscat.getFullYear() + '-' +
    String(muscat.getMonth()+1).padStart(2,'0') + '-' +
    String(muscat.getDate()).padStart(2,'0');
  var lastVisit = localStorage.getItem('_lastVisit');

  if(lastVisit !== todayStr) {{
    fetch(SUPA_URL + '/rest/v1/page_views', {{
      method: 'POST',
      headers: {{
        'apikey': SUPA_KEY,
        'Authorization': 'Bearer ' + SUPA_KEY,
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal'
      }},
      body: JSON.stringify({{ page: 'export', visitor_id: visitorId }})
    }}).then(function() {{
      localStorage.setItem('_lastVisit', todayStr);
    }}).catch(function(){{}});
  }}
}})();
</script>

<!-- ══════ LONG-PRESS SCREENSHOT ══════ -->
<script>
(function(){{
  // ── overlay للتحميل ──
  var overlay = document.createElement('div');
  overlay.id = 'ssOverlay';
  overlay.style.cssText = 'display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:99999;align-items:center;justify-content:center;flex-direction:column;gap:12px;';
  overlay.innerHTML = '<div style="width:44px;height:44px;border:4px solid rgba(255,255,255,.3);border-top-color:#fff;border-radius:50%;animation:ssSpin .8s linear infinite;"></div><span style="color:#fff;font-size:14px;font-weight:700;">جاري الالتقاط…</span><style>@keyframes ssSpin{{to{{transform:rotate(360deg)}}}}</style>';
  document.body.appendChild(overlay);

  function showOverlay(){{ overlay.style.display='flex'; }}
  function hideOverlay(){{ overlay.style.display='none'; }}

  // ── مشاركة واتساب ──
  function shareToWhatsApp(canvas, label) {{
    canvas.toBlob(function(blob) {{
      hideOverlay();
      if (!blob) {{ alert('فشل إنشاء الصورة'); return; }}
      // حاول Web Share API (موبايل)
      if (navigator.canShare && navigator.canShare({{ files: [new File([blob],'roster.png',{{type:'image/png'}})] }})) {{
        navigator.share({{
          title: label,
          text: label,
          files: [new File([blob], 'roster.png', {{type:'image/png'}})]
        }}).catch(function(){{}});
      }} else {{
        // fallback: تنزيل الصورة
        var url = URL.createObjectURL(blob);
        var a = document.createElement('a');
        a.href = url; a.download = 'roster.png'; a.click();
        setTimeout(function(){{ URL.revokeObjectURL(url); }}, 3000);
      }}
    }}, 'image/png');
  }}

  // ── التقاط عنصر مع فتح كل الـ details أولاً ──
  function captureElement(el, label, onlyOpenDetails) {{
    showOverlay();
    // احفظ حالة details الحالية وافتحها كلها
    var detailsEls = el.querySelectorAll('details.shiftCard');
    var wasOpen = [];
    detailsEls.forEach(function(d, i) {{
      wasOpen[i] = d.hasAttribute('open');
      d.setAttribute('open', '');
    }});
    // انتظر لحظة للرسم
    setTimeout(function() {{
      html2canvas(el, {{
        scale: 2,
        useCORS: true,
        backgroundColor: '#eef1f7',
        logging: false,
        removeContainer: true
      }}).then(function(canvas) {{
        // أعد الحالة الأصلية
        detailsEls.forEach(function(d, i) {{
          if (!wasOpen[i]) d.removeAttribute('open');
        }});
        shareToWhatsApp(canvas, label);
      }}).catch(function() {{
        detailsEls.forEach(function(d, i) {{
          if (!wasOpen[i]) d.removeAttribute('open');
        }});
        hideOverlay();
        alert('فشل الالتقاط، حاول مرة أخرى');
      }});
    }}, 300);
  }}

  // ── التقاط مناوبة واحدة ──
  function captureShift(summaryEl, shiftLabel) {{
    showOverlay();
    var detailsEl = summaryEl.closest('details.shiftCard');
    if (!detailsEl) {{ hideOverlay(); return; }}
    var wasOpen = detailsEl.hasAttribute('open');
    detailsEl.setAttribute('open', '');
    setTimeout(function() {{
      html2canvas(detailsEl, {{
        scale: 2,
        useCORS: true,
        backgroundColor: '#fff',
        logging: false,
        removeContainer: true
      }}).then(function(canvas) {{
        if (!wasOpen) detailsEl.removeAttribute('open');
        shareToWhatsApp(canvas, shiftLabel);
      }}).catch(function() {{
        if (!wasOpen) detailsEl.removeAttribute('open');
        hideOverlay();
        alert('فشل الالتقاط');
      }});
    }}, 200);
  }}

  // ── long-press helper ──
  function addLongPress(el, callback, ms) {{
    ms = ms || 700;
    var timer = null;
    var moved = false;

    function start(e) {{
      moved = false;
      timer = setTimeout(function() {{
        if (!moved) {{
          e.preventDefault && e.preventDefault();
          callback(e);
        }}
      }}, ms);
    }}
    function cancel() {{ clearTimeout(timer); }}
    function onMove() {{ moved = true; clearTimeout(timer); }}

    var longPressed = false;
    el.addEventListener('touchstart', function(e) {{
      longPressed = false;
      moved = false;
      timer = setTimeout(function() {{
        if (!moved) {{
          longPressed = true;
          callback(e);
        }}
      }}, ms);
    }}, {{passive: true}});
    el.addEventListener('touchend', function(e) {{
      if (timer) {{ clearTimeout(timer); timer = null; }}
      // لا نعيد تعيين longPressed هنا — ننتظر حتى يصل حدث click
    }});
    el.addEventListener('touchmove', onMove, {{passive: true}});
    el.addEventListener('click', function(e) {{
      if (longPressed) {{
        e.preventDefault();
        e.stopPropagation();
        e.stopImmediatePropagation();
        longPressed = false;
        return false;
      }}
    }}, true);
    el.addEventListener('mousedown', start);
    el.addEventListener('mouseup', cancel);
    el.addEventListener('mousemove', onMove);
    el.addEventListener('contextmenu', function(e) {{ e.preventDefault(); }});
    el.style.webkitTouchCallout = 'none';
    el.style.webkitUserSelect = 'none';
    el.style.userSelect = 'none';
    el.style.touchAction = 'manipulation';
  }}

  // ── ربط deptHead: ضغط مطوّل → التقط القسم كله ──
  document.querySelectorAll('.deptHead').forEach(function(head) {{
    var deptName = head.getAttribute('data-dept') || 'Roster';
    addLongPress(head, function() {{
      var card = head.closest('.deptCard');
      if (card) captureElement(card, deptName);
    }});
  }});

  // ── ربط shiftSummary: ضغط مطوّل → التقط المناوبة فقط ──
  document.querySelectorAll('.shiftSummary').forEach(function(summary) {{
    var shiftKey = summary.getAttribute('data-shift-key') || '';
    addLongPress(summary, function() {{
      captureShift(summary, shiftKey);
    }});
  }});

}})();
</script>

<script src="/roster-site/eid-overlay.js"></script>
<script src="/roster-site/absence-alert.js"></script>
</body>
</html>"""


def generate_date_pages_for_month(wb, year: int, month: int, pages_base: str, source_name: str = "", min_date: str = "", max_date: str = ""):
    """
    Generate static pages for each day of the given month.
    Used by the date picker to navigate to different dates.

    If wb is None, it still generates pages but shows a 'no roster' notice.
    """
    import calendar
    from datetime import datetime as dt

    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        try:
            date_obj = dt(year, month, day, tzinfo=TZ)
            dow = (date_obj.weekday() + 1) % 7  # Sun=0
            active_group = current_shift_key(dt.now(TZ))

            dept_cards_all = []
            dept_cards_now = []
            employees_total_all = 0
            employees_total_now = 0
            depts_count = 0

            notice_html = ""
            if wb is None:
                notice_html = (
                    "<div class='deptCard' style='padding:14px;border:1px dashed rgba(15,23,42,.20);background:#fff;'>"
                    "⚠️ لا يوجد روستر لهذا الشهر بعد.</div>"
                )
            else:
                for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
                    if sheet_name not in wb.sheetnames:
                        continue

                    ws = wb[sheet_name]
                    days_row, date_row = find_days_and_dates_rows(ws)
                    day_col = find_day_col(ws, days_row, date_row, dow, day)

                    if not (days_row and date_row and day_col):
                        continue

                    start_row = date_row + 1
                    emp_col = find_employee_col(ws, start_row=start_row)
                    daynum_to_col = get_daynum_to_col(ws, date_row)
                    if not emp_col:
                        continue

                    buckets = {k: [] for k in GROUP_ORDER}
                    buckets_now = {k: [] for k in GROUP_ORDER}

                    for r in range(start_row, ws.max_row + 1):
                        name = norm(ws.cell(row=r, column=emp_col).value)
                        if not looks_like_employee_name(name):
                            continue

                        daynum_to_raw = {dn: norm(ws.cell(row=r, column=col).value) for dn, col in daynum_to_col.items()}
                        raw = daynum_to_raw.get(day, "")
                        if not looks_like_shift_code(raw):
                            continue

                        label, grp = map_shift(raw)

                        up = norm(raw).upper()
                        if grp == "Annual Leave":
                            if up == "AL" or "ANNUAL LEAVE" in up or up == "LV":
                                suf = range_suffix_for_day(day, daynum_to_raw, "AL")
                                if suf:
                                    label = suf
                        elif grp == "Sick Leave":
                            if up == "SL" or "SICK LEAVE" in up:
                                suf = range_suffix_for_day(day, daynum_to_raw, "SL")
                                if suf:
                                    label = suf
                        elif grp == "Training":
                            if up == "TR" or "TRAINING" in up:
                                suf = range_suffix_for_day(day, daynum_to_raw, "TR")
                                if suf:
                                    label = suf

                        buckets.setdefault(grp, []).append({"name": name, "shift": label})

                        # /now page: include ALL groups so the shift filter buttons work for any date
                        buckets_now.setdefault(grp, []).append({"name": name, "shift": label})
                    dept_color = UNASSIGNED_COLOR if dept_name == "Unassigned" else DEPT_COLORS[idx % len(DEPT_COLORS)]
                    open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None

                    dept_cards_all.append(dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full))
                    dept_cards_now.append(dept_card_html(dept_name, dept_color, buckets_now, open_group=active_group))

                    employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
                    employees_total_now += sum(len(buckets_now.get(g, [])) for g in GROUP_ORDER)
                    depts_count += 1

                if employees_total_all == 0:
                    notice_html = (
                        "<div class='deptCard' style='padding:14px;border:1px dashed rgba(15,23,42,.20);background:#fff;'>"
                        "ℹ️ لا توجد بيانات لهذا التاريخ في الروستر.</div>"
                    )

            try:
                date_label = date_obj.strftime("%-d %B %Y")
            except Exception:
                date_label = date_obj.strftime("%d %B %Y")

            iso_date = date_obj.strftime("%Y-%m-%d")
            sent_time = date_obj.strftime("%H:%M")
            last_updated = date_obj.strftime("%d%b%Y / %H:%M").upper()

            full_url = f"{pages_base}/"
            now_url = f"{pages_base}/now/"

            html_full = page_shell_html(
                date_label=date_label,
                iso_date=iso_date,
                employees_total=employees_total_all,
                departments_total=depts_count,
                dept_cards_html="\n".join(dept_cards_all),
                cta_url=now_url,
                sent_time=sent_time,
                source_name=source_name,
                last_updated=last_updated,
                is_now_page=False,
                min_date=min_date,
                max_date=max_date,
                notice_html=notice_html,
            )

            html_now = page_shell_html(
                date_label=date_label,
                iso_date=iso_date,
                employees_total=employees_total_now,
                departments_total=depts_count,
                dept_cards_html="\n".join(dept_cards_now),
                cta_url=full_url,
                sent_time=sent_time,
                source_name=source_name,
                last_updated=last_updated,
                is_now_page=True,
                min_date=min_date,
                max_date=max_date,
                notice_html=notice_html,
            )

            date_dir = f"docs/date/{iso_date}"
            os.makedirs(date_dir, exist_ok=True)
            os.makedirs(f"{date_dir}/now", exist_ok=True)

            with open(f"{date_dir}/index.html", "w", encoding="utf-8") as f:
                f.write(html_full)

            with open(f"{date_dir}/now/index.html", "w", encoding="utf-8") as f:
                f.write(html_now)

        except Exception as e:
            print(f"Skipping {year}-{month:02d}-{day:02d}: {e}")
            continue


# =========================
# Email
# =========================
def get_subscriber_emails():
    """
    يقرأ قائمة الإيميلات من Google Apps Script
    """
    subscriber_url = os.environ.get('SUBSCRIBE_URL', '').strip()
    
    if not subscriber_url:
        return os.environ.get('MAIL_TO', '').strip()
    
    try:
        print(f"📥 Fetching subscriber emails...")
        response = requests.get(subscriber_url, timeout=10)
        response.raise_for_status()
        
        email_list = response.text.strip()
        
        if not email_list:
            print("⚠️ No subscribers found, using MAIL_TO")
            return os.environ.get('MAIL_TO', '').strip()
        
        subscriber_count = len([e for e in email_list.split(',') if e.strip()])
        print(f"✅ Found {subscriber_count} active subscribers")
        
        return email_list
        
    except Exception as e:
        print(f"❌ Error fetching subscribers: {e}")
        print("⚠️ Falling back to MAIL_TO")
        return os.environ.get('MAIL_TO', '').strip()


def send_email(subject: str, html: str):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASS and MAIL_FROM):
        return

    recipient_list = get_subscriber_emails()
    recipients = [x.strip() for x in recipient_list.split(",") if x.strip()]

    if not recipients:
        print("⚠️ No recipients found, skipping email")
        return

    msg = MIMEText(html, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = MAIL_FROM
    msg["To"] = MAIL_TO or MAIL_FROM  # ✅ لا تعرض قائمة المشتركين

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(MAIL_FROM, recipients, msg.as_string())
        print(f"✅ Sent to {len(recipients)} subscribers")
    except smtplib.SMTPException as e:
        print(f"❌ SMTP error while sending email: {e}")
    except OSError as e:
        print(f"❌ Network error while connecting to SMTP server: {e}")


def build_pretty_email_html(active_shift_key: str, now: datetime, all_shifts_by_dept: list, pages_base: str) -> str:
    """
    Builds a beautifully formatted HTML email showing ALL shifts for the day.
    The active shift is highlighted with a ⚡ badge.
    all_shifts_by_dept = [{"dept": ..., "shifts": {"Morning": [...], "Afternoon": [...], ...}}, ...]
    """
    # ✅ عرض جميع المناوبات في الإيميل
    include_groups = GROUP_ORDER

    def get_group_employees(shifts_data: dict, group_key: str):
        # ✅ إرجاع جميع الموظفين بدون فلترة
        return shifts_data.get(group_key, []) or []

    # Calculate totals across included groups only
    total_employees = 0
    depts_with_employees = 0

    for d in all_shifts_by_dept:
        shifts_data = d.get("shifts", {})
        dept_total = 0
        for g in include_groups:
            dept_total += len(get_group_employees(shifts_data, g))
        if dept_total > 0:
            depts_with_employees += 1
            total_employees += dept_total

    # Determine current shift colors for header
    shift_colors = SHIFT_COLORS.get(active_shift_key, SHIFT_COLORS["Other"])
    shift_icon = shift_colors.get("icon", "⏰")

    # Build department cards with ALL shifts
    dept_cards = []
    for idx, d in enumerate(all_shifts_by_dept):
        dept_name = d["dept"]
        shifts_data = d["shifts"]
        # Skip if department has no employees for the included groups
        dept_total = 0
        for g in include_groups:
            dept_total += len(get_group_employees(shifts_data, g))
        if dept_total == 0:
            continue
        # Determine department color
        if dept_name == "Unassigned":
            dept_color = UNASSIGNED_COLOR
        else:
            dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]
        # Build shift sections (only active shift + matching Standby)
        shift_sections = ""
        for group_key in GROUP_ORDER:
            if group_key not in include_groups:
                continue
            employees = get_group_employees(shifts_data, group_key)
            if not employees:
                continue

            # Get shift display name
            shift_display_names = {
                "Morning": "Morning",
                "Afternoon": "Afternoon",
                "Night": "Night",
                "Off Day": "Off Day",
                "Annual Leave": "Annual Leave",   # ✅ Fixed: was "Leave" (wrong key)
                "Sick Leave": "Sick Leave",        # ✅ Added missing key
                "Training": "Training",
                "Standby": "Standby",
                "Other": "Other"
            }
            display_name = shift_display_names.get(group_key, group_key)
            
            colors = SHIFT_COLORS.get(group_key, SHIFT_COLORS["Other"])
            count = len(employees)

            # Highlight active shift
            is_active = (group_key == active_shift_key)
            active_border = f"border:2px solid {colors['border']};" if is_active else f"border:1px solid {colors['border']};"
            active_badge = "⚡" if is_active else ""

            # Build employee rows
            rows_html = ""
            for i, e in enumerate(employees):
                bg_color = "rgba(15,23,42,.03)" if i % 2 == 1 else "transparent"
                rows_html += f"""
                    <tr>
                      <td style="padding:10px 14px;border-top:1px solid rgba(15,23,42,.06);background:{bg_color};">
                        <span style="font-size:14px;font-weight:700;color:#1e293b;">{e['name']}</span>
                      </td>
                      <td style="padding:10px 14px;border-top:1px solid rgba(15,23,42,.06);text-align:right;background:{bg_color};">
                        <span style="font-size:13px;font-weight:600;color:{colors['status_color']};white-space:nowrap;">{e['shift']}</span>
                      </td>
                    </tr>"""

            shift_sections += f"""
              <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:10px;background:{colors['bg']};border-radius:12px;overflow:hidden;{active_border}">
                <!-- Shift Header -->
                <tr>
                  <td colspan="2" style="padding:10px 14px;background:{colors['summary_bg']};border-bottom:1px solid {colors['summary_border']};">
                    <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                      <tr>
                        <td style="padding:0;">
                          <span style="font-size:18px;margin-right:8px;">{colors['icon']}</span>
                          <span style="font-size:15px;font-weight:800;color:{colors['label_color']};letter-spacing:-.1px;">{display_name} {active_badge}</span>
                        </td>
                        <td style="text-align:right;padding:0;">
                          <span style="display:inline-block;padding:4px 12px;border-radius:20px;background:{colors['count_bg']};color:{colors['count_color']};font-size:13px;font-weight:800;">{count}</span>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
                <!-- Employees -->
                {rows_html}
              </table>"""

        # Department icon SVG
        icon_svg = """<svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>"""

        dept_cards.append(f"""
          <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:18px;background:#fff;border-radius:18px;overflow:hidden;border:1px solid rgba(15,23,42,.07);box-shadow:0 4px 18px rgba(15,23,42,.08);">
            <!-- Colored top gradient bar -->
            <tr>
              <td colspan="2" style="height:5px;background:linear-gradient(to right,{dept_color['grad_from']},{dept_color['grad_to']});padding:0;"></td>
            </tr>
            
            <!-- Department Header -->
            <tr>
              <td colspan="2" style="padding:14px 16px;border-bottom:2px solid {dept_color['border']};background:#fff;">
                <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                  <tr>
                    <td style="width:46px;padding:0;">
                      <div style="width:44px;height:44px;border-radius:12px;background:{dept_color['light']};color:{dept_color['base']};display:flex;align-items:center;justify-content:center;">
                        {icon_svg}
                      </div>
                    </td>
                    <td style="padding:0 0 0 12px;">
                      <span style="font-size:18px;font-weight:800;color:#1e293b;letter-spacing:-.2px;display:block;">{dept_name}</span>
                    </td>
                    <td style="text-align:right;padding:0;">
                      <div style="display:inline-block;min-width:52px;padding:8px 12px;border-radius:12px;background:{dept_color['light']};border:1px solid {dept_color['border']};text-align:center;">
                        <span style="font-size:10px;opacity:.7;display:block;text-transform:uppercase;letter-spacing:.5px;color:{dept_color['base']};margin-bottom:1px;">Total</span>
                        <span style="font-size:17px;font-weight:900;color:{dept_color['base']};display:block;">{dept_total}</span>
                      </div>
                    </td>
                  </tr>
                </table>
              </td>
            </tr>

            <!-- Included Shifts -->
            <tr>
              <td colspan="2" style="padding:10px;">
                {shift_sections}
              </td>
            </tr>
          </table>
        """)

    dept_html = "".join(dept_cards)
    sent_time = now.strftime("%H:%M")
    date_str = now.strftime("%d %B %Y")
    last_updated = now.strftime("%d%b%Y / %H:%M").upper()

    # Translate active_shift_key display
    shift_display_map = {
        "Morning": "Morning Shift",
        "Afternoon": "Afternoon Shift", 
        "Night": "Night Shift"
    }
    shift_display = shift_display_map.get(active_shift_key, active_shift_key)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster - {date_str}</title>
  <style>
    @media only screen and (max-width: 600px) {{
      .mobile-padding {{ padding: 12px !important; }}
      .mobile-font {{ font-size: 13px !important; }}
      .header-icon {{ font-size: 56px !important; }}
    }}
  </style>
</head>
<body style="margin:0;padding:0;background:#eef1f7;font-family:'Segoe UI',system-ui,-apple-system,BlinkMacSystemFont,Roboto,Helvetica,Arial,sans-serif;-webkit-font-smoothing:antialiased;">
  
  <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#eef1f7;">
    <tr>
      <td align="center" style="padding:20px 14px;">
        
        <!-- Main Container -->
        <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="max-width:680px;width:100%;margin:0 auto;">
          
          <!-- Compact Header with Gradient -->
          <tr>
  <td style="padding:0;">
    <table role="presentation" cellpadding="0" cellspacing="0" border="0"
      style="width:100%;
             background:linear-gradient(135deg,#1e40af 0%,#1976d2 50%,#0ea5e9 100%);
             border-radius:20px 20px 0 0;
             overflow:hidden;
             box-shadow:0 8px 26px rgba(30,64,175,.25);
             position:relative;">
      <tr>
        <td style="padding:18px 18px;text-align:center;position:relative;">

          <!-- Decorative circles -->
          <div style="position:absolute;top:-50px;right:-60px;width:160px;height:160px;border-radius:50%;background:rgba(255,255,255,.08);"></div>
          <div style="position:absolute;bottom:-70px;left:-50px;width:180px;height:180px;border-radius:50%;background:rgba(255,255,255,.06);"></div>

          <!-- Icon -->
          <div class="header-icon"
               style="font-size:40px;margin-bottom:6px;position:relative;z-index:1;">📋</div>

          <!-- Title -->
          <h1 style="margin:0;
                     font-size:22px;
                     font-weight:800;
                     color:#ffffff;
                     letter-spacing:-.4px;
                     position:relative;
                     z-index:1;">
            Duty Roster
          </h1>

          <!-- Active Shift -->
          <div style="margin-top:8px;
                      display:inline-block;
                      background:rgba(255,255,255,.22);
                      padding:6px 16px;
                      border-radius:18px;
                      font-size:13px;
                      font-weight:700;
                      color:#ffffff;
                      letter-spacing:.3px;
                      position:relative;
                      z-index:1;">
            {shift_icon} {shift_display}
          </div>

          <!-- Date -->
          <div style="margin-top:6px;
                      display:inline-block;
                      background:rgba(255,255,255,.16);
                      padding:5px 14px;
                      border-radius:16px;
                      font-size:12px;
                      font-weight:600;
                      color:#ffffff;
                      letter-spacing:.2px;
                      position:relative;
                      z-index:1;">
            📅 {date_str}
          </div>

        </td>
      </tr>
    </table>
  </td>
</tr>

          <!-- Summary Stats -->
          <tr>
            <td style="padding:0 14px;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;margin-top:18px;">
                <tr>
                  <td style="width:50%;padding-right:6px;">
                    <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border:1px solid rgba(15,23,42,.10);border-radius:16px;box-shadow:0 3px 12px rgba(15,23,42,.07);">
                      <tr>
                        <td style="padding:16px;text-align:center;">
                          <div style="font-size:28px;font-weight:900;color:#1e40af;margin-bottom:4px;">{total_employees}</div>
                          <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;">Employees</div>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <td style="width:50%;padding-left:6px;">
                    <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border:1px solid rgba(15,23,42,.10);border-radius:16px;box-shadow:0 3px 12px rgba(15,23,42,.07);">
                      <tr>
                        <td style="padding:16px;text-align:center;">
                          <div style="font-size:28px;font-weight:900;color:#059669;margin-bottom:4px;">{depts_with_employees}</div>
                          <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;">Departments</div>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Department Cards with ALL Shifts -->
          <tr>
            <td style="padding:0 14px;">
              {dept_html}
            </td>
          </tr>

          <!-- Call to Action Buttons -->
          <tr>
            <td style="padding:22px 14px;text-align:center;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
                <tr>
                  <td style="padding:0 7px 0 0;">
                    <a href="{pages_base}/now/" style="display:inline-block;padding:15px 30px;border-radius:16px;background:linear-gradient(135deg,#1e40af,#1976d2);color:#ffffff;text-decoration:none;font-weight:800;font-size:15px;box-shadow:0 6px 22px rgba(30,64,175,.35);white-space:nowrap;">
                      🔄 Refresh Now
                    </a>
                  </td>
                  <td style="padding:0 0 0 7px;">
                    <a href="{pages_base}/" style="display:inline-block;padding:15px 30px;border-radius:16px;background:linear-gradient(135deg,#0ea5e9,#06b6d4);color:#ffffff;text-decoration:none;font-weight:800;font-size:15px;box-shadow:0 6px 22px rgba(14,165,233,.35);white-space:nowrap;">
                      📋 View Full Roster
                    </a>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="padding:0 14px 22px;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border-radius:0 0 20px 20px;border:1px solid rgba(15,23,42,.08);border-top:none;">
                <tr>
                  <td style="padding:18px;text-align:center;color:#94a3b8;font-size:13px;line-height:1.9;">
                    <strong style="color:#475569;">Last Updated:</strong> <strong style="color:#1e40af;">{last_updated}</strong>
                    <br>
                    Total on duty: <strong style="color:#64748b;">{total_employees} employees</strong> across <strong style="color:#64748b;">{depts_with_employees} departments</strong>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

        </table>
        
      </td>
    </tr>
  </table>

</body>
</html>"""




# =========================
# Main
# =========================
def main():
    if not EXCEL_URL:
        raise RuntimeError("EXCEL_URL missing")

    parser = argparse.ArgumentParser(description='Generate roster pages and send email')
    parser.add_argument('--date', help='Override roster date (YYYY-MM-DD)')
    args = parser.parse_args()

    now = datetime.now(TZ)
    if args.date:
        try:
            y, m, d = [int(x) for x in args.date.strip().split('-')]
            now = datetime(y, m, d, now.hour, now.minute, tzinfo=TZ)
        except Exception:
            raise RuntimeError('Invalid --date format. Use YYYY-MM-DD')

    today_dow = (now.weekday() + 1) % 7
    today_day = now.day
    active_group = current_shift_key(now)

    # pages_base - cleanup
    pages_base_raw = PAGES_BASE_URL or infer_pages_base_url()
    pages_base = pages_base_raw.rstrip("/")
    if pages_base.endswith("/now"):
        pages_base = pages_base[:-4]

    # ─────────────────────────────────────────────────────────────
    # FIX #1: تحميل Excel - عند الفشل نكمل بالكاش (لا نخرج)
    # ─────────────────────────────────────────────────────────────
    data = None
    try:
        data = download_excel(EXCEL_URL)
        print("✅ Excel downloaded successfully")
    except Exception as e:
        print(f"WARNING: Could not download Excel: {e}")
        print("Will attempt to use cached rosters...")

    # ─────────────────────────────────────────────────────────────
    # FIX #2: كل هذا الكود الآن خارج الـ except - يعمل دائماً
    # ─────────────────────────────────────────────────────────────

    # قراءة اسم الملف واستخراج الشهر
    source_name = get_source_name()
    incoming_key = month_key_from_filename(source_name) if source_name else None
    print(f"📄 Source file: {source_name}")
    print(f"📅 Detected month: {incoming_key or 'unknown'}")

    # FIX #3: حفظ في الكاش فقط إذا نجح التحميل
    if data and incoming_key:
        xlsx_path, meta_path = cache_paths(incoming_key)
        try:
            write_bytes(xlsx_path, data)
            write_json(meta_path, {
                "month_key": incoming_key,
                "original_filename": source_name,
                "downloaded_at": datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S %Z"),
            })
            print(f"✅ Cached roster: {xlsx_path}")
        except Exception as e:
            print(f"WARNING: failed caching roster: {e}")
    elif not incoming_key:
        print("⚠️ Could not detect month from filename; cache skipped for this run.")

    # حساب الأشهر الثلاثة
    prev_y, prev_m = add_months(now.year, now.month, -1)
    next_y, next_m = add_months(now.year, now.month, +1)

    prev_key = f"{prev_y:04d}-{prev_m:02d}"
    curr_key = f"{now.year:04d}-{now.month:02d}"
    next_key = f"{next_y:04d}-{next_m:02d}"

    # نطاق الـ date picker: من أول الشهر السابق إلى آخر الشهر القادم
    min_date = f"{prev_y:04d}-{prev_m:02d}-01"
    max_date = f"{next_y:04d}-{next_m:02d}-{calendar.monthrange(next_y, next_m)[1]:02d}"

    print(f"📅 Month range: {prev_key} → {curr_key} → {next_key}")

    # تحميل الكاش لكل شهر
    wb_prev = try_load_cached_workbook(prev_key)
    wb_curr = try_load_cached_workbook(curr_key)
    wb_next = try_load_cached_workbook(next_key)

    # FIX #4: استخدام البيانات المحملة لتعبئة الكاش الناقص - workbook واحد فقط
    if data:
        wb_data = load_workbook(BytesIO(data), data_only=True)
        if wb_prev is None and incoming_key == prev_key:
            wb_prev = wb_data
            print(f"✅ Using downloaded data for {prev_key}")
        elif wb_curr is None and incoming_key == curr_key:
            wb_curr = wb_data
            print(f"✅ Using downloaded data for {curr_key}")
        elif wb_next is None and incoming_key == next_key:
            wb_next = wb_data
            print(f"✅ Using downloaded data for {next_key}")

    print(f"📦 Cache status: prev={'✅' if wb_prev else '❌'} | curr={'✅' if wb_curr else '❌'} | next={'✅' if wb_next else '❌'}")

    # توليد صفحات الأشهر الثلاثة
    generate_date_pages_for_month(
        wb_prev, prev_y, prev_m, pages_base,
        source_name=cached_source_name(prev_key) or source_name,
        min_date=min_date, max_date=max_date
    )
    generate_date_pages_for_month(
        wb_curr, now.year, now.month, pages_base,
        source_name=cached_source_name(curr_key) or source_name,
        min_date=min_date, max_date=max_date
    )
    generate_date_pages_for_month(
        wb_next, next_y, next_m, pages_base,
        source_name=cached_source_name(next_key) or source_name,
        min_date=min_date, max_date=max_date
    )

    # الصفحة الرئيسية تستخدم الشهر الحالي
    wb = wb_curr

    # ─────────────────────────────────────────────────────────────
    # من هنا: توليد الصفحة الرئيسية docs/index.html و docs/now/
    # ─────────────────────────────────────────────────────────────
    if wb is None:
        os.makedirs("docs", exist_ok=True)
        os.makedirs("docs/now", exist_ok=True)

        try:
            date_label = now.strftime("%-d %B %Y")
        except Exception:
            date_label = now.strftime("%d %B %Y")

        iso_date = now.strftime("%Y-%m-%d")
        last_updated = now.strftime("%d%b%Y / %H:%M").upper()

        notice_html = (
            "<div class='deptCard' style='padding:14px;border:1px dashed rgba(15,23,42,.20);background:#fff;'>"
            "⚠️ لا يوجد روستر للشهر الحالي محفوظ بعد. (قد يكون الروستر الجديد للشهر القادم وصل مبكرًا)</div>"
        )

        html_full = page_shell_html(
            date_label=date_label,
            iso_date=iso_date,
            employees_total=0,
            departments_total=0,
            dept_cards_html="",
            cta_url=f"{pages_base}/now/",
            sent_time=now.strftime("%H:%M"),
            source_name=cached_source_name(curr_key) or source_name,
            last_updated=last_updated,
            is_now_page=False,
            min_date=min_date,
            max_date=max_date,
            notice_html=notice_html,
        )

        html_now = page_shell_html(
            date_label=date_label,
            iso_date=iso_date,
            employees_total=0,
            departments_total=0,
            dept_cards_html="",
            cta_url=f"{pages_base}/",
            sent_time=now.strftime("%H:%M"),
            source_name=cached_source_name(curr_key) or source_name,
            last_updated=last_updated,
            is_now_page=True,
            min_date=min_date,
            max_date=max_date,
            notice_html=notice_html,
        )

        with open("docs/index.html", "w", encoding="utf-8") as f:
            f.write(html_full)

        with open("docs/now/index.html", "w", encoding="utf-8") as f:
            f.write(html_now)

        print("⚠️ Skipping email (no current-month roster workbook).")
        return


    dept_cards_all = []
    dept_cards_now = []
    all_shifts_by_dept = []
    employees_total_all = 0
    employees_total_now = 0
    depts_count = 0

    for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        day_col = find_day_col(ws, days_row, date_row, today_dow, today_day)

        if not (days_row and date_row and day_col):
            continue

        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        daynum_to_col = get_daynum_to_col(ws, date_row)
        if not emp_col:
            continue

        buckets = {k: [] for k in GROUP_ORDER}
        buckets_now = {k: [] for k in GROUP_ORDER}

        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue

            daynum_to_raw = {dn: norm(ws.cell(row=r, column=col).value) for dn, col in daynum_to_col.items()}

            raw = daynum_to_raw.get(today_day, "")
            if not looks_like_shift_code(raw):
                continue

            label, grp = map_shift(raw)

            up = norm(raw).upper()
            # إضافة نطاق التواريخ FROM TO للإجازات السنوية
            if grp == "Annual Leave":
                if up == "AL" or "ANNUAL LEAVE" in up or up == "LV":
                    suf = range_suffix_for_day(today_day, daynum_to_raw, "AL")
                    if suf:
                        label = suf  # فقط النطاق الزمني بدون اسم الإجازة
            # إضافة نطاق التواريخ FROM TO للإجازات المرضية
            elif grp == "Sick Leave":
                if up == "SL" or "SICK LEAVE" in up:
                    suf = range_suffix_for_day(today_day, daynum_to_raw, "SL")
                    if suf:
                        label = suf  # فقط النطاق الزمني بدون اسم الإجازة
            # إضافة نطاق التواريخ FROM TO للتدريب
            elif grp == "Training":
                if up == "TR" or "TRAINING" in up:
                    suf = range_suffix_for_day(today_day, daynum_to_raw, "TR")
                    if suf:
                        label = suf  # فقط النطاق الزمني بدون اسم الإجازة
            
            buckets.setdefault(grp, []).append({"name": name, "shift": label})

            if grp == active_group:
                buckets_now.setdefault(grp, []).append({"name": name, "shift": label})

        all_shifts_by_dept.append({"dept": dept_name, "shifts": buckets})

        if dept_name == "Unassigned":
            dept_color = UNASSIGNED_COLOR
        else:
            dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]

        open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None
        card_all = dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full)
        dept_cards_all.append(card_all)

        # صفحة /now/ تحتوي على كل الورديات (سيتم الفلترة بـ JavaScript)
        card_now = dept_card_html(dept_name, dept_color, buckets, open_group=active_group)
        dept_cards_now.append(card_now)

        employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
        # حساب فقط الوردية الحالية لـ total في /now/
        employees_total_now += sum(len(buckets.get(g, [])) for g in [active_group, "Off Day", "Annual Leave", "Sick Leave", "Training", "Standby", "Other"])

        depts_count += 1

    os.makedirs("docs", exist_ok=True)
    os.makedirs("docs/now", exist_ok=True)

    try:
        date_label = now.strftime("%-d %B %Y")
    except Exception:
        date_label = now.strftime("%d %B %Y")

    iso_date = now.strftime("%Y-%m-%d")
    sent_time = now.strftime("%H:%M")
    last_updated = now.strftime("%d%b%Y / %H:%M").upper()

    full_url = f"{pages_base}/"
    now_url = f"{pages_base}/now/"

    html_full = page_shell_html(
        date_label=date_label,
        iso_date=iso_date,
        employees_total=employees_total_all,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_all),
        cta_url=now_url,
        sent_time=sent_time,
        source_name=cached_source_name(curr_key) or source_name,
        last_updated=last_updated,
        min_date=min_date,
        max_date=max_date,
        is_now_page=False,
    )
    html_now = page_shell_html(
        date_label=date_label,
        iso_date=iso_date,
        employees_total=employees_total_now,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_now),
        cta_url=full_url,
        sent_time=sent_time,
        source_name=cached_source_name(curr_key) or source_name,
        last_updated=last_updated,
        min_date=min_date,
        max_date=max_date,
        is_now_page=True,
    )

    with open("docs/index.html", "w", encoding="utf-8") as f:
        f.write(html_full)

    with open("docs/now/index.html", "w", encoding="utf-8") as f:
        f.write(html_now)

    # Write source name for my-schedules page to display
    _src = cached_source_name(curr_key) or source_name
    if _src:
        os.makedirs("docs/my-schedules", exist_ok=True)
        with open("docs/my-schedules/source.txt", "w", encoding="utf-8") as f:
            f.write(_src)

    # Email: يُرسل فقط عندما يقرر الـ YML ذلك (SEND_EMAIL=true)
    # الحالتان: الساعة 7 صباحاً أو عند تغيير الروستر
    should_send_email = os.environ.get("SEND_EMAIL", "false").strip().lower() == "true"
    if should_send_email:
        changed    = os.environ.get("CHANGED", "false").strip().lower() == "true"
        is_morning = os.environ.get("IS_MORNING_EMAIL", "false").strip().lower() == "true"
        if changed:
            subject = f"🔄 Duty Roster Updated — {now.strftime('%d %B %Y')}"
        elif is_morning:
            subject = f"🌅 Duty Roster — صباح الخير {now.strftime('%d %B %Y')}"
        else:
            subject = f"📋 Duty Roster — {now.strftime('%d %B %Y')}"
        print(f"✅ Sending email — changed={changed}, morning={is_morning}")
        email_html = build_pretty_email_html(active_group, now, all_shifts_by_dept, pages_base)
        send_email(subject, email_html)
    else:
        print("⏭️  Skipping email — SEND_EMAIL=false")


if __name__ == "__main__":
    main()
