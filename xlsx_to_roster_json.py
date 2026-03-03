"""
xlsx_to_roster_json.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
يقرأ ملفات الروستر من مجلد rosters/ (xlsx)
ويولّد docs/data/roster.json بالصيغة التي يتوقعها offload_monitor.py

الصيغة المطلوبة:
{
  "days": {
    "2026-03-03": {
      "cards_html": "<div class='deptCard'>...</div>"
    }
  }
}

cards_html يحتوي على:
  - div.deptCard لكل قسم
  - div.deptTitle للاسم
  - div.shiftCard لكل وردية
  - div.shiftLabel لاسم الوردية
  - div.empRow + div.empName لكل موظف
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations

import os
import re
import json
import calendar
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

# ══════════════════════════════════════════════════════════════
#  الإعدادات
# ══════════════════════════════════════════════════════════════

TZ          = ZoneInfo("Asia/Muscat")
ROSTERS_DIR = Path(os.environ.get("ROSTERS_DIR", "rosters"))
OUTPUT_PATH = Path("docs/data/roster.json")

DEPARTMENTS = [
    "Officers",
    "Supervisors",
    "Load Control",
    "Export Checker",
    "Export Operators",
    "Unassigned",
]

DAYS = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

SHIFT_MAP = {
    "MN06": "Morning",  "ME06": "Morning",  "ME07": "Morning",
    "MN12": "Afternoon","AN13": "Afternoon","AE14": "Afternoon",
    "NN21": "Night",    "NE22": "Night",
}

GROUP_ORDER = [
    "Morning", "Afternoon", "Night",
    "Standby", "Off Day", "Annual Leave",
    "Sick Leave", "Training", "Other",
]

# ══════════════════════════════════════════════════════════════
#  دوال مساعدة (مستنسخة من generate_and_send_22_.py)
# ══════════════════════════════════════════════════════════════

def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00A0", " ")).strip()

def to_western_digits(s: str) -> str:
    arabic = {'٠':'0','١':'1','٢':'2','٣':'3','٤':'4','٥':'5','٦':'6','٧':'7','٨':'8','٩':'9'}
    farsi  = {'۰':'0','۱':'1','۲':'2','۳':'3','۴':'4','۵':'5','۶':'6','۷':'7','۸':'8','۹':'9'}
    mp = {**arabic, **farsi}
    return "".join(mp.get(ch, ch) for ch in str(s or ""))

def norm(s) -> str:
    return clean(to_western_digits(s))

def looks_like_time(s: str) -> bool:
    up = norm(s).upper()
    return bool(
        re.match(r"^\d{3,4}\s*H?\s*-\s*\d{3,4}\s*H?$", up)
        or re.match(r"^\d{3,4}\s*H$", up)
        or re.match(r"^\d{3,4}$", up)
    )

def looks_like_employee_name(s: str) -> bool:
    v = norm(s)
    if not v:
        return False
    up = v.upper()
    if looks_like_time(up):
        return False
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", up):
        return False
    if re.search(r"-\s*\d{3,}", v) and re.search(r"[A-Za-z\u0600-\u06FF]", v):
        return True
    parts = [p for p in v.split(" ") if p]
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", v) and len(parts) >= 2)

def looks_like_shift_code(s: str) -> bool:
    v = norm(s).upper()
    if not v:
        return False
    if looks_like_time(v):
        return False
    if v in ["OFF","O","LV","TR","ST","SL","AL","STM","STN","STNE22","STME06","STMN06","STAE14","OT"]:
        return True
    if re.match(r"^(MN|AN|NN|NT|ME|AE|NE)\d{1,2}", v):
        return True
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", v):
        return True
    if len(v) >= 3 and re.search(r"[A-Z]", v):
        return True
    return False

def map_shift(code: str) -> tuple[str, str]:
    c0 = norm(code)
    c  = c0.upper()
    if not c or c == "0":
        return ("-", "Other")
    if c == "AL" or c == "LV" or "ANNUAL LEAVE" in c:
        return ("AL", "Annual Leave")
    if c == "SL" or "SICK LEAVE" in c:
        return ("SL", "Sick Leave")
    if c in ["TR"] or "TRAINING" in c:
        return ("TR", "Training")
    if c in ["ST","STM","STN","STNE22","STME06","STMN06","STAE14"] or "STANDBY" in c:
        return (c0, "Standby")
    if c == "OT" or c.startswith("OT"):
        return (c0, "Standby")
    if c in ["OFF","O"] or re.search(r"(REST|OFF\s*DAY|REST\/OFF)", c):
        return ("OFF", "Off Day")
    if c in SHIFT_MAP:
        return SHIFT_MAP[c], SHIFT_MAP[c]
    return (c0, "Other")

def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False

def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    return sum(1 for d in DAYS if any(d in x for x in ups))

def find_days_and_dates_rows(ws, scan_rows: int = 80):
    max_r    = min(ws.max_row, scan_rows)
    days_row = None
    for r in range(1, max_r + 1):
        if _count_day_tokens(_row_values(ws, r)) >= 3:
            days_row = r
            break
    if not days_row:
        return None, None
    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        nums = sum(1 for v in _row_values(ws, r) if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break
    return days_row, date_row

def get_daynum_to_col(ws, date_row: int) -> dict:
    m = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(v):
            m[int(float(v))] = c
    return m

def find_employee_col(ws, start_row: int):
    for c in range(1, min(ws.max_column, 15) + 1):
        found = sum(
            1 for r in range(start_row, min(start_row + 20, ws.max_row) + 1)
            if looks_like_employee_name(norm(ws.cell(row=r, column=c).value))
        )
        if found >= 3:
            return c
    return None

# ══════════════════════════════════════════════════════════════
#  استخراج موظفي يوم معين من ورقة Excel
# ══════════════════════════════════════════════════════════════

def extract_day(ws, day_num: int) -> dict[str, list]:
    """يرجع dict: {group_name: [{"name": str, "shift": str}, ...]}"""
    days_row, date_row = find_days_and_dates_rows(ws)
    if not (days_row and date_row):
        return {}

    daynum_to_col = get_daynum_to_col(ws, date_row)
    if day_num not in daynum_to_col:
        return {}

    start_row = date_row + 1
    emp_col   = find_employee_col(ws, start_row=start_row)
    if not emp_col:
        return {}

    buckets: dict[str, list] = {k: [] for k in GROUP_ORDER}

    for r in range(start_row, ws.max_row + 1):
        name = norm(ws.cell(row=r, column=emp_col).value)
        if not looks_like_employee_name(name):
            continue
        raw = norm(ws.cell(row=r, column=daynum_to_col[day_num]).value)
        if not looks_like_shift_code(raw):
            continue
        label, grp = map_shift(raw)
        buckets.setdefault(grp, []).append({"name": name, "shift": label})

    return {k: v for k, v in buckets.items() if v}

# ══════════════════════════════════════════════════════════════
#  توليد cards_html لكل يوم
# ══════════════════════════════════════════════════════════════

def build_cards_html(wb, day_num: int) -> str:
    """يولّد HTML بالصيغة التي يتوقعها offload_monitor.py"""
    html_parts = []

    for sheet_name in DEPARTMENTS:
        if sheet_name not in wb.sheetnames:
            continue

        ws      = wb[sheet_name]
        buckets = extract_day(ws, day_num)
        if not buckets:
            continue

        shift_cards = ""
        for grp in GROUP_ORDER:
            emps = buckets.get(grp, [])
            if not emps:
                continue
            emp_rows = ""
            for e in emps:
                emp_rows += (
                    f"<div class='empRow'>"
                    f"<div class='empName'>{e['name']}</div>"
                    f"<div class='empShift'>{e['shift']}</div>"
                    f"</div>"
                )
            shift_cards += (
                f"<div class='shiftCard'>"
                f"<div class='shiftLabel'>{grp}</div>"
                f"{emp_rows}"
                f"</div>"
            )

        if not shift_cards:
            continue

        html_parts.append(
            f"<div class='deptCard'>"
            f"<div class='deptTitle'>{sheet_name}</div>"
            f"{shift_cards}"
            f"</div>"
        )

    return "\n".join(html_parts)

# ══════════════════════════════════════════════════════════════
#  تحميل workbook من الكاش
# ══════════════════════════════════════════════════════════════

def load_wb(month_key: str):
    xlsx_path = ROSTERS_DIR / f"{month_key}.xlsx"
    if not xlsx_path.exists():
        return None
    try:
        with open(xlsx_path, "rb") as f:
            return load_workbook(BytesIO(f.read()), data_only=True)
    except Exception as e:
        print(f"  [xlsx_to_json] Failed to load {xlsx_path}: {e}")
        return None

def add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    y, m = year, month + delta
    while m <= 0:
        y -= 1; m += 12
    while m > 12:
        y += 1; m -= 12
    return y, m

# ══════════════════════════════════════════════════════════════
#  نقطة الدخول
# ══════════════════════════════════════════════════════════════

def main():
    now = datetime.now(TZ)
    print(f"[xlsx_to_roster_json] Starting — {now.strftime('%Y-%m-%d %H:%M')} ...")

    # نبني بيانات الشهر السابق + الحالي + القادم
    months = []
    for delta in (-1, 0, 1):
        y, m = add_months(now.year, now.month, delta)
        months.append((y, m, f"{y:04d}-{m:02d}"))

    days_dict: dict[str, dict] = {}

    for year, month, month_key in months:
        wb = load_wb(month_key)
        if wb is None:
            print(f"  [xlsx_to_roster_json] No roster found for {month_key} — skipping.")
            continue

        print(f"  [xlsx_to_roster_json] Processing {month_key} ...")
        days_in_month = calendar.monthrange(year, month)[1]

        for day_num in range(1, days_in_month + 1):
            iso_date   = date(year, month, day_num).isoformat()
            cards_html = build_cards_html(wb, day_num)

            if cards_html:
                days_dict[iso_date] = {"cards_html": cards_html}
                print(f"    ✅ {iso_date} — data added")
            else:
                print(f"    ⚠️  {iso_date} — no employee data found")

    output = {"generated_at": now.isoformat(), "days": days_dict}

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(output, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"[xlsx_to_roster_json] Done. ✅  ({len(days_dict)} days written → {OUTPUT_PATH})")


if __name__ == "__main__":
    main()
